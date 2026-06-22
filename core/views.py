from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, update_session_auth_hash, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.db.models import Q, Count
from .models import Student, Profile, Faculty, Specialty, AuditLog, ExcelDownloadRequest, Notification, Announcement, ChatMessage, MessageReaction, SiteLogo, LoginPageContent, StudentPayment, Application, LandingPageContent, LandingNews, Comment, ExcelImportConfig
from .forms import StudentForm, CommentForm, PasswordChangeForm, FacultyForm, SpecialtyForm, ExcelRequestForm, ExcelAdminForm, AnnouncementForm, ChatMessageForm, MessageReactionForm, ExcelUploadForm, UserRegistrationForm, StudentRegistrationForm, EmployeeRegistrationForm, ApplicationForm, ApplicationReviewForm, UsernameChangeForm, SiteLogoForm, LoginPageContentForm, ProfileEditForm, ExcelImportConfigForm
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
from django.http import HttpResponse, JsonResponse
from datetime import datetime
import bleach # XSS sanitizatsiyasi uchun
import os

# --- Yordamchi fayl yuklash funksiyasi ---
def _validate_file_upload(file_data, allowed_extensions, max_size_mb=5):
    if file_data:
        ext = os.path.splitext(file_data.name)[1].lower()
        if ext not in allowed_extensions:
            raise ValueError(f"Nomaqbul fayl formati! Faqat quyidagilarga ruxsat beriladi: {", ".join(allowed_extensions)}")
        
        max_size_bytes = max_size_mb * 1024 * 1024
        if file_data.size > max_size_bytes:
            raise ValueError(f"Fayl hajmi {max_size_mb} MB dan oshmasligi kerak!")
    return file_data



from io import BytesIO
from django.utils import timezone
from rest_framework.authtoken.models import Token
from django.contrib.auth.models import User
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from axes.decorators import axes_dispatch

# --- Yordamchi funksiyalar ---
def is_super_admin(user):
    return hasattr(user, 'profile') and user.profile.has_role('super_admin')

def is_admin(user):
    return hasattr(user, 'profile') and user.profile.has_role('admin')

def is_operator(user):
    return hasattr(user, 'profile') and user.profile.has_role('operator')

def is_applicant(user):
    return hasattr(user, 'profile') and user.profile.has_role('applicant')

def is_student(user):
    """Talaba (qara ro'yhat) tekshiruvi"""
    return hasattr(user, 'profile') and user.profile.has_role('student')

def is_employee(user):
    """Xodim (qara ro'yhat) tekshiruvi"""
    return hasattr(user, 'profile') and user.profile.has_role('employee')

def is_regular_user(user):
    """Ariza beruvchi, talaba yoki xodim tekshiruvi"""
    return is_applicant(user) or is_student(user) or is_employee(user)

def log_action(user, action, details=""):
    AuditLog.objects.create(user=user, action=action, details=details)

# --- Landing Page (Bosh sahifa) ---
def landing_page(request):
    """BUI uslubidagi landing page - tizimga kirishdan oldin"""
    return redirect('login')
    # Avtomatik hisoblangan statistika
    default_stats = {
        'total_students': Student.objects.filter(status='approved').count(),
        'total_applications': Application.objects.count(),
        'faculties_count': Faculty.objects.count(),
        'specialties_count': Specialty.objects.count(),
    }

    # Landing content ni olish va override larni tekshirish
    landing_content = LandingPageContent.get_active()
    
    # Override larni avtomatik son ustiga QO'SHISH (buzish emas)
    stats = default_stats.copy()
    if landing_content:
        if landing_content.override_students_count is not None:
            stats['total_students'] += int(landing_content.override_students_count)
        if landing_content.override_applications_count is not None:
            stats['total_applications'] += int(landing_content.override_applications_count)
        if landing_content.override_faculties_count is not None:
            stats['faculties_count'] += int(landing_content.override_faculties_count)
        if landing_content.override_specialties_count is not None:
            stats['specialties_count'] += int(landing_content.override_specialties_count)

    faculties = Faculty.objects.all()[:6]
    news_items = LandingNews.objects.filter(is_active=True)[:6]

    context = {
        'stats': stats,
        'faculties': faculties,
        'landing_content': landing_content,
        'news_items': news_items,
    }
    return render(request, 'landing.html', context)


# --- Landing Page boshqaruvi (Super Admin) ---
@login_required
def landing_settings(request):
    """Super admin uchun landing page sozlamalari"""
    messages.error(request, "Bosh sahifa sozlamalari vaqtinchalik faolsizlantirilgan!")
    return redirect('dashboard')

    content = LandingPageContent.get_active()

    if request.method == 'POST':
        if not content:
            content = LandingPageContent()

        content.hero_title = bleach.clean(request.POST.get('hero_title', content.hero_title if content.pk else ''), tags=[], attributes={}, strip=True)
        content.hero_subtitle = request.POST.get('hero_subtitle', '')
        content.hero_description = request.POST.get('hero_description', '')
        content.contact_address = request.POST.get('contact_address', '')
        content.contact_phone = request.POST.get('contact_phone', '')
        content.contact_email = request.POST.get('contact_email', '')
        content.contact_location_url = request.POST.get('contact_location_url', '')
        content.about_title = request.POST.get('about_title', '')
        content.about_text = request.POST.get('about_text', '')
        
        # Statistika override lari
        content.override_students_count = request.POST.get('override_students_count') if request.POST.get('override_students_count') else None
        content.override_applications_count = request.POST.get('override_applications_count') if request.POST.get('override_applications_count') else None
        content.override_faculties_count = request.POST.get('override_faculties_count') if request.POST.get('override_faculties_count') else None
        content.override_specialties_count = request.POST.get('override_specialties_count') if request.POST.get('override_specialties_count') else None
        
        content.is_active = True

        # Rasm/video o'chirish
        if request.POST.get('delete_hero_image') == '1' and content.hero_image:
            content.hero_image.delete(save=False)
            content.hero_image = None
        if request.POST.get('delete_hero_video') == '1' and content.hero_video:
            content.hero_video.delete(save=False)
            content.hero_video = None

        # Yangi fayl yuklash
        if request.FILES.get('hero_image'):
            content.hero_image = request.FILES['hero_image']
        if request.FILES.get('hero_video'):
            content.hero_video = request.FILES['hero_video']

        content.save()
        log_action(request.user, "Landing page sozlamalari o'zgartirildi")
        messages.success(request, "Landing page sozlamalari saqlandi!")
        return redirect('landing_settings')

    # Joriy statistikani hisoblash
    current_stats = {
        'total_students': Student.objects.filter(status='approved').count(),
        'total_applications': Application.objects.count(),
        'faculties_count': Faculty.objects.count(),
        'specialties_count': Specialty.objects.count(),
    }
    
    news_items = LandingNews.objects.all()
    return render(request, 'landing_settings.html', {
        'content': content,
        'news_items': news_items,
        'stats': current_stats,
    })


@login_required
def landing_news_add(request):
    """Landing page uchun yangilik qo'shish"""
    active_role = request.session.get('active_role', request.user.profile.role)
    if active_role != 'super_admin':
        messages.error(request, "Ruxsat yo'q.")
        return redirect('dashboard')

    if request.method == 'POST':
        news = LandingNews(
            title=request.POST.get('title'),
            description=request.POST.get('description'),
            video_url=request.POST.get('video_url', ''),
            link=request.POST.get('link', ''),
            order=int(request.POST.get('order') or 0),
            is_active=True,
        )
        if request.FILES.get('image'):
            news.image = request.FILES['image']
        news.save()
        log_action(request.user, "Landing yangilik qo'shildi", news.title)
        messages.success(request, "Yangilik qo'shildi!")

    return redirect('landing_settings')


@login_required
def landing_news_delete(request, pk):
    """Landing yangilikni o'chirish"""
    active_role = request.session.get('active_role', request.user.profile.role)
    if active_role != 'super_admin':
        messages.error(request, "Ruxsat yo'q.")
        return redirect('dashboard')

    news = get_object_or_404(LandingNews, pk=pk)
    title = news.title
    news.delete()
    log_action(request.user, "Landing yangilik o'chirildi", title)
    messages.success(request, "Yangilik o'chirildi.")
    return redirect('landing_settings')


@login_required
def landing_news_edit(request, pk):
    """Landing yangilikni tahrirlash"""
    active_role = request.session.get('active_role', request.user.profile.role)
    if active_role != 'super_admin':
        messages.error(request, "Ruxsat yo'q.")
        return redirect('dashboard')

    news = get_object_or_404(LandingNews, pk=pk)
    if request.method == 'POST':
        news.title = request.POST.get('title', news.title)
        news.description = request.POST.get('description', news.description)
        news.video_url = request.POST.get('video_url', '')
        news.link = request.POST.get('link', '')
        news.order = int(request.POST.get('order') or 0)
        news.is_active = request.POST.get('is_active') == 'on'
        if request.FILES.get('image'):
            news.image = request.FILES['image']
        news.save()
        messages.success(request, "Yangilik yangilandi!")
        return redirect('landing_settings')

    return render(request, 'landing_news_edit.html', {'news': news})

def offline_page(request):
    """Offline sahifa - service worker fallback"""
    return render(request, 'offline.html')


# --- Public Faculty and Specialty Detail Pages ---
def faculty_detail(request, faculty_id):
    """Fakultet detail sahifasi (omaviy)"""
    faculty = get_object_or_404(Faculty, id=faculty_id)
    specialties = faculty.specialty_set.all()
    
    # Fakultet haqida qo'shimcha ma'lumotlar
    context = {
        'faculty': faculty,
        'specialties': specialties,
        'specialties_count': specialties.count(),
        'students_count': Student.objects.filter(specialty__faculty=faculty, status='approved').count(),
        'applications_count': Student.objects.filter(specialty__faculty=faculty, status='pending').count(),
    }
    return render(request, 'faculty_detail.html', context)


def specialty_detail(request, specialty_id):
    """Mutaxasislik detail sahifasi (omaviy)"""
    specialty = get_object_or_404(Specialty, id=specialty_id)
    
    # Mutaxasislik haqida qo'shimcha ma'lumotlar
    context = {
        'specialty': specialty,
        'faculty': specialty.faculty,
        'students_count': Student.objects.filter(specialty=specialty, status='approved').count(),
        'applications_count': Student.objects.filter(specialty=specialty, status='pending').count(),
        'contract_amount': specialty.contract_amount,
    }
    return render(request, 'specialty_detail.html', context)


def all_faculties(request):
    """Barcha fakultetlar ro'yxati sahifasi"""
    faculties = Faculty.objects.all()
    
    context = {
        'faculties': faculties,
        'faculties_count': faculties.count(),
        'total_specialties': Specialty.objects.count(),
        'total_students': Student.objects.filter(status='approved').count(),
    }
    return render(request, 'all_faculties.html', context)


def all_specialties(request):
    """Barcha mutaxasisliklar ro'yxati sahifasi"""
    specialties = Specialty.objects.all().select_related('faculty')
    
    context = {
        'specialties': specialties,
        'specialties_count': specialties.count(),
        'total_faculties': Faculty.objects.count(),
        'total_students': Student.objects.filter(status='approved').count(),
    }
    return render(request, 'all_specialties.html', context)


@login_required
def offline_queue_page(request):
    """Offline navbat - kutmoqdagi yozuvlar"""
    return render(request, 'offline_queue.html')


def get_active_logo_url(request):
    """Faol logotip URL ni olish va session ga saqlash"""
    active_logo = SiteLogo.get_active_logo()
    if active_logo:
        request.session['site_logo'] = active_logo.logo.url
        return active_logo.logo.url
    else:
        request.session.pop('site_logo', None)
        return None

def get_filtered_students(user, active_role=None):
    students = Student.objects.all()
    # Agar active_role berilmagan bo'lsa, user.profile.role dan foydalanamiz
    role = active_role if active_role else (user.profile.role if hasattr(user, 'profile') else 'operator')
    
    if role == 'super_admin':
        pass  # Super admin barcha talabalarni ko'radi
    elif role == 'admin':
        # Admin biriktirilgan operatori bo'lsa faqat o'sha operatorlarning talabalarini ko'radi
        # Bo'lmaganda barcha talabalarni ko'radi
        subordinate_count = user.subordinate_operators.count()
        if subordinate_count > 0:
            operator_ids = list(user.subordinate_operators.values_list('user_id', flat=True))
            # Admin o'z operatorlari qo'shgan talabalarni + mustaqil qo'shilgan talabalarni ko'radi
            students = students.filter(Q(created_by__in=operator_ids) | Q(created_by__isnull=True))
        # Aks holda barcha talabalarni ko'radi
    elif role == 'operator':
        # Operator o'zining barcha talabalarini + mustaqil qo'shilgan talabalarni ko'radi
        students = students.filter(Q(created_by=user) | Q(created_by__isnull=True))
    elif role == 'applicant':
        students = Student.objects.none()  # Ariza beruvchi talabalar bo'limini ko'ra olmaydi
    elif role == 'student':
        students = Student.objects.none()  # Talaba (ariza beruvchi) talabalar bo'limini ko'ra olmaydi
    elif role == 'employee':
        students = Student.objects.none()  # Xodim (ariza beruvchi) talabalar bo'limini ko'ra olmaydi
    elif role == 'hemis_id_adder':
        students = students.filter(status='approved')  # HEMIS ID qo'shuvchi faqat tasdiqlangan talabalarni ko'radi
    elif role == 'talaba_viewer':
        students = students.filter(status__in=['approved', 'rejected'])  # Talabani ko'ruvchi tasdiqlangan va rad etilgan talabalarni ko'radi
    elif role == 'archive':
        # Archive faqat tasdiqlangan talabalarni ko'radi, kutilmoqda emas
        students = students.filter(status='approved')
    elif role == 'accountant':
        students = students.filter(status='approved')  # Accountant faqat tasdiqlangan talabalarni ko'radi
    
    return students

def send_notification(recipient, title, message, notification_type='general', sender=None, related_object_id=None):
    """Xabar yuborish"""
    notification = Notification.objects.create(
        recipient=recipient,
        sender=sender,
        title=title,
        message=message,
        notification_type=notification_type,
        related_object_id=related_object_id
    )
    return notification

def get_unread_chat_count(user):
    """O'qilmagan chat xabarlar soni"""
    return ChatMessage.objects.filter(recipient=user, is_read=False).count()

# --- Autentifikatsiya (O'z bloklash logikasi bilan) ---
def login_view(request):
    from .captcha import MathCaptcha
    from django.db import OperationalError
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        captcha_answer = request.POST.get('captcha_answer')
        captcha_correct = request.POST.get('captcha_correct')
        
        # Captcha tekshirish
        if not MathCaptcha.verify_answer(captcha_answer, captcha_correct):
            messages.error(request, 'Matematik misol noto\'g\'ri javob berildi!')
            # Yangi captcha yaratish
            captcha_data = MathCaptcha.generate_captcha()
            try:
                login_page_content = LoginPageContent.get_active_content()
            except OperationalError:
                login_page_content = None
            return render(request, 'login.html', {
                'captcha': captcha_data,
                'unread_chat_count': get_unread_chat_count(request.user) if request.user.is_authenticated else 0,
                'username': username,  # Login ni saqlab qolish
                'login_page_content': login_page_content,
            })
        
        # O'z bloklash logikasi
        from axes.models import AccessAttempt
        from django.utils import timezone
        from datetime import timedelta
        
        # Super admin uchun bloklarni o'chirish
        try:
            user_obj = User.objects.get(username=username)
            if hasattr(user_obj, 'profile') and user_obj.profile.role == 'super_admin':
                AccessAttempt.objects.filter(username=username).delete()
        except User.DoesNotExist:
            pass
        
        # User bloklanganligini tekshirish - faqat username bo'yicha
        if username:
            cutoff_time = timezone.now() - timedelta(minutes=15)  # 15 daqiqa
            # Faqat username bo'yicha sanash, IP va user_agent e'tiborga olinmaydi
            failed_attempts = AccessAttempt.objects.filter(
                username=username,
                attempt_time__gte=cutoff_time
            ).count()
            
            if failed_attempts >= 5:
                return render(request, 'account_lockout.html')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # XAVFSIZLIK: Session ID ni qayta generatsiya qilish (session fixation himoyasi)
            from django.contrib.sessions.models import Session
            # Eski sessiyani o'chirish
            if request.session.session_key:
                try:
                    old_session = Session.objects.get(session_key=request.session.session_key)
                    old_session.delete()
                except Session.DoesNotExist:
                    pass
            
            # Yangi sessiya yaratish
            request.session.cycle_key()
            
            login(request, user)
            # Muvaffaqiyatli login da bloklarni o'chirish
            AccessAttempt.objects.filter(username=username).delete()
            log_action(user, 'Tizimga kirdi', f'IP: {request.META.get("REMOTE_ADDR")}')
            messages.success(request, f"Xush kelibsiz, {username}!")
            return redirect('dashboard')
        else:
            # Xatolik qayd etish - faqat username bo'yicha sanash
            # IP va user_agent e'tiborga olinmaydi, shunda bir qurilma bloklanmaydi
            existing_attempt = AccessAttempt.objects.filter(
                username=username,
                attempt_time__gte=timezone.now() - timedelta(minutes=15)
            ).first()
            
            if existing_attempt:
                # Mavjud urinishni yangilash
                existing_attempt.attempt_time = timezone.now()
                existing_attempt.failures_since_start += 1
                existing_attempt.ip_address = request.META.get("REMOTE_ADDR")
                existing_attempt.user_agent = request.META.get("HTTP_USER_AGENT", "")
                existing_attempt.save()
            else:
                # Yangi urinish yaratish
                AccessAttempt.objects.create(
                    username=username,
                    ip_address=request.META.get("REMOTE_ADDR"),
                    user_agent=request.META.get("HTTP_USER_AGENT", ""),
                    attempt_time=timezone.now(),
                    failures_since_start=1
                )
            
            log_action(None, 'Muvaffaqiyatsiz kirish urinishi', 
                      f'Username: {username}, IP: {request.META.get("REMOTE_ADDR")}')
            messages.error(request, 'Login yoki parol noto\'g\'ri')
    
    # GET request yoki POST xatolik bo'lsa - yangi captcha
    captcha_data = MathCaptcha.generate_captcha()
    
    try:
        login_page_content = LoginPageContent.get_active_content()
    except OperationalError:
        login_page_content = None
    
    return render(request, 'login.html', {
        'captcha': captcha_data,
        'unread_chat_count': get_unread_chat_count(request.user) if request.user.is_authenticated else 0,
        'login_page_content': login_page_content,
    })

def logout_view(request):
    if request.user.is_authenticated:
        log_action(request.user, 'Tizimdan chiqdi')
    logout(request)
    messages.info(request, 'Tizimdan chiqdingiz')
    return redirect('login')

def _send_verification_code(email, code):
    """Email ga tasdiqlash kodini yuborish - EmailSettings (parolni unutdim bilan bir xil) orqali"""
    from .models import EmailSettings
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    email_settings = EmailSettings.get_active_settings()
    if not email_settings:
        raise Exception("Faol email sozlama topilmadi! Super admin Email Sozlamalari bo'limidan SMTP qo'shsin.")

    server = smtplib.SMTP(email_settings.smtp_host, email_settings.smtp_port)
    server.ehlo()
    if email_settings.use_tls:
        server.starttls()
        server.ehlo()

    clean_email = email_settings.email.replace('\u200b', '').strip()
    
    # Get password from settings or environment if not present in email_settings
    raw_password = getattr(email_settings, 'password', None)
    if not raw_password:
        import os
        from django.conf import settings
        raw_password = os.environ.get('EMAIL_HOST_PASSWORD') or getattr(settings, 'EMAIL_HOST_PASSWORD', '')
        
    clean_password = (raw_password or '').replace('\u200b', '').strip()
    server.login(clean_email, clean_password)

    msg = MIMEMultipart()
    msg['From'] = f"Test-sinov-ucun <{email_settings.email}>"
    msg['To'] = email
    msg['Reply-To'] = email_settings.email
    msg['Subject'] = 'Test-sinov-ucun - Tasdiqlash Kodi'

    body = (
        f"Assalomu alaykum!\n\n"
        f"Sizning ro'yxatdan o'tish kodingiz: {code}\n\n"
        f"Bu kod 10 daqiqa davomida yaroqli.\n\n"
        f"Agar siz ro'yxatdan o'tishni so'ramagan bo'lsangiz, ushbu xabarni e'tiborsiz qoldiring.\n\n"
        f"Hurmat bilan"
    )
    msg.attach(MIMEText(body, 'plain', 'utf-8'))

    server.send_message(msg)
    server.quit()
    print(f"Verification email yuborildi: {email_settings.email} -> {email}")


def register_student_view(request):
    """Talaba ro'yxatdan o'tish - Email tasdiqlash bilan (2 bosqich)"""
    messages.warning(request, "Talaba sifatida ro'yxatdan o'tish vaqtincha to'xtatilgan.")
    return redirect('login')

    import random
    from django.utils import timezone as dj_tz
    from datetime import timedelta

    # === BOSQICH 2: Tasdiqlash kodini tekshirish ===
    if request.method == 'POST' and request.POST.get('action') == 'verify_code':
        pending = request.session.get('pending_registration')
        if not pending:
            messages.error(request, "Sessiya muddati tugadi. Iltimos, qaytadan ro'yxatdan o'ting.")
            return redirect('register_student')

        # Muddat tekshirish
        expires_at = pending.get('expires_at', 0)
        if dj_tz.now().timestamp() > expires_at:
            request.session.pop('pending_registration', None)
            messages.error(request, "Tasdiqlash kodi muddati tugagan. Qaytadan urinib ko'ring.")
            return redirect('register_student')

        entered_code = request.POST.get('verification_code', '').strip()
        if entered_code != pending.get('code'):
            messages.error(request, "Kod noto'g'ri! Iltimos, qaytadan urinib ko'ring.")
            return render(request, 'register_verify_email.html', {
                'email': pending.get('form_data', {}).get('email'),
            })

        # Kod to'g'ri - foydalanuvchini yaratish
        form = StudentRegistrationForm(pending['form_data'])
        if form.is_valid():
            try:
                user = form.save()
                request.session.pop('pending_registration', None)
                login(request, user)
                log_action(user, "Ro'yxatdan o'tdi (email tasdiqlangan)", f'Role: student')
                messages.success(request, "Email tasdiqlandi! Ro'yxatdan o'tdingiz. Arizangiz ko'rib chiqilmoqda.")
                return redirect('my_application')
            except Exception as e:
                messages.error(request, f"Xatolik: {str(e)}")
                return redirect('register_student')
        else:
            messages.error(request, "Forma ma'lumotlari noto'g'ri.")
            return redirect('register_student')

    # === Kodni qaytadan yuborish ===
    if request.method == 'POST' and request.POST.get('action') == 'resend_code':
        pending = request.session.get('pending_registration')
        if pending:
            new_code = f"{random.randint(100000, 999999)}"
            pending['code'] = new_code
            pending['expires_at'] = (dj_tz.now() + timedelta(minutes=10)).timestamp()
            request.session['pending_registration'] = pending
            try:
                _send_verification_code(pending['form_data']['email'], new_code)
                messages.success(request, "Yangi kod yuborildi!")
            except Exception as e:
                messages.error(request, f"Email yuborishda xatolik: {str(e)}")
            return render(request, 'register_verify_email.html', {
                'email': pending['form_data']['email'],
            })
        return redirect('register_student')

    # === BOSQICH 1: Forma yuborish va kod yuborish ===
    if request.method == 'POST':
        form = StudentRegistrationForm(request.POST)
        if form.is_valid():
            # Kod yaratish va session ga saqlash
            code = f"{random.randint(100000, 999999)}"
            email = form.cleaned_data['email']

            request.session['pending_registration'] = {
                'form_data': {k: str(v) for k, v in request.POST.items() if k != 'csrfmiddlewaretoken'},
                'code': code,
                'expires_at': (dj_tz.now() + timedelta(minutes=10)).timestamp(),
            }

            try:
                _send_verification_code(email, code)
                messages.info(request, f"Tasdiqlash kodi {email} ga yuborildi!")
                return render(request, 'register_verify_email.html', {'email': email})
            except Exception as e:
                request.session.pop('pending_registration', None)
                messages.error(request, f"Email yuborishda xatolik: {str(e)}. Iltimos, email manzilingizni tekshiring.")
        else:
            messages.error(request, "Iltimos, forma maydonlarini to'g'ri to'ldiring.")
    else:
        form = StudentRegistrationForm()

    return render(request, 'register_student.html', {
        'form': form,
        'unread_chat_count': 0,
    })

def register_employee_view(request):
    """Xodim ro'yxatdan o'tish sahifasi - Rol olish uchun"""
    if request.method == 'POST':
        form = EmployeeRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            log_action(user, 'Ro\'yxatdan o\'tdi', f'Role: employee (Xodim - Rol olish uchun)')
            messages.success(request, f"Ro'yxatdan o'tdingiz! Arizangiz admin va super adminni ko'rib chiqishini kutaymiz.")
            return redirect('my_application')
    else:
        form = EmployeeRegistrationForm()

    return render(request, 'register_employee.html', {
        'form': form,
        'unread_chat_count': 0,
    })

@login_required
@axes_dispatch
def submit_application(request):
    """Redirect to my_application - ro'yxatdan o'tishda avtomatik Application yaratiladi"""
    return redirect('my_application')

@login_required
@axes_dispatch
def my_application(request):
    """Foydalanuvchi o'z arizasini ko'rishi"""
    if not is_regular_user(request.user):
        messages.error(request, "Sizda bu sahifaga kirish huquqi yo'q.")
        return redirect('dashboard')
    
    application = Application.objects.filter(user=request.user).order_by('-submitted_at').first()
    
    if not application:
        # Agar ariza yo'q bo'lsa, yangi ariza sahifasiga yo'nalt
        return redirect('submit_application')
    
    return render(request, 'my_application.html', {
        'application': application,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

@login_required
@axes_dispatch
def applications_list(request):
    """Arizalar bo'limi - FAQAT saytdan o'zi mustaqil ro'yxatdan o'tganlar ko'rinadi"""
    messages.warning(request, "Arizalar bo'limi vaqtincha to'xtatilgan.")
    return redirect('dashboard')

    active_role = request.session.get('active_role', request.user.profile.role)
    if active_role not in ['admin', 'super_admin']:
        messages.error(request, "Sizda bu bo'limni ko'rish huquqi yo'q.")
        return redirect('dashboard')

    # Ariza turini olish (student yoki employee)
    application_type = request.GET.get('type', 'all')  # 'all', 'student', 'employee'

    # Operator/admin tomonidan qo'shilgan talabalar pasportlari (created_by mavjud)
    operator_added_passports = list(
        Student.objects.exclude(created_by__isnull=True).values_list('passport', flat=True)
    )

    # Faqat o'zi mustaqil ro'yxatdan o'tganlar (Student.created_by IS NULL yoki Student yo'q)
    applications = Application.objects.select_related('user', 'faculty', 'specialty', 'reviewed_by') \
        .exclude(passport__in=operator_added_passports)

    # Student va Employee ni alohida filtrlash
    # Student: faculty va specialty mavjud, pasport bo'sh emas
    # Employee: faculty va specialty None, pasport bo'sh
    if application_type == 'student':
        applications = applications.filter(faculty__isnull=False, specialty__isnull=False)
    elif application_type == 'employee':
        applications = applications.filter(faculty__isnull=True, specialty__isnull=True)

    status_filter = request.GET.get('status')
    search = request.GET.get('search')
    if status_filter:
        applications = applications.filter(status=status_filter)
    if search:
        applications = applications.filter(
            Q(full_name__icontains=search) |
            Q(passport__icontains=search) |
            Q(email__icontains=search)
        )

    # Statistika - faqat o'zi ro'yxatdan o'tganlar uchun
    self_registered_qs = Application.objects.exclude(passport__in=operator_added_passports)
    total_count = applications.count()
    student_count = self_registered_qs.filter(faculty__isnull=False, specialty__isnull=False).count()
    employee_count = self_registered_qs.filter(faculty__isnull=True, specialty__isnull=True).count()

    # Tartib va pagination
    applications = applications.order_by('-submitted_at')
    from django.core.paginator import Paginator
    per_page = int(request.GET.get('per_page', 25))
    per_page = min(max(per_page, 10), 200)
    paginator = Paginator(applications, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'applications_list.html', {
        'applications': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'status_filter': status_filter,
        'search': search,
        'application_type': application_type,
        'total_count': total_count,
        'student_count': student_count,
        'employee_count': employee_count,
        'unread_chat_count': get_unread_chat_count(request.user),
    })


@login_required
@axes_dispatch
def all_applicants_list(request):
    """Barcha foydalanuvchilar - Operator qo'shgan + O'zi ro'yxatdan o'tgan (admin/super_admin uchun)"""
    active_role = request.session.get('active_role', request.user.profile.role)
    if active_role not in ['admin', 'super_admin']:
        messages.error(request, "Sizda bu bo'limni ko'rish huquqi yo'q.")
        return redirect('dashboard')

    source_filter = request.GET.get('source', 'all')  # 'all', 'operator', 'self'
    status_filter = request.GET.get('status')
    search = request.GET.get('search')

    # Manba bo'yicha ajratish: Student.created_by mavjud bo'lsa - operator qo'shgan
    # created_by IS NULL bo'lsa - o'zi ro'yxatdan o'tgan
    operator_students = Student.objects.exclude(created_by__isnull=True) \
        .select_related('faculty', 'specialty', 'created_by', 'approved_by')
    self_students = Student.objects.filter(created_by__isnull=True) \
        .select_related('faculty', 'specialty', 'approved_by')

    if status_filter:
        operator_students = operator_students.filter(status=status_filter)
        self_students = self_students.filter(status=status_filter)

    if search:
        operator_students = operator_students.filter(
            Q(full_name__icontains=search) |
            Q(passport__icontains=search) |
            Q(phone__icontains=search)
        )
        self_students = self_students.filter(
            Q(full_name__icontains=search) |
            Q(passport__icontains=search) |
            Q(phone__icontains=search)
        )

    combined = []

    if source_filter in ('all', 'operator'):
        for s in operator_students:
            added_by_user = s.created_by
            full = f"{added_by_user.first_name} {added_by_user.last_name}".strip()
            added_by_name = full if full else added_by_user.username
            combined.append({
                'id': s.id,
                'full_name': s.full_name,
                'passport': s.passport,
                'phone': s.phone,
                'faculty': s.faculty.name if s.faculty else '-',
                'specialty': s.specialty.name if s.specialty else '-',
                'status': s.status,
                'status_display': s.get_status_display(),
                'submitted_at': s.created_at,
                'source': 'operator',
                'source_display': 'Operator qo\'shgan',
                'added_by': added_by_name,
                'detail_url': f'/process/{s.id}/',
            })

    if source_filter in ('all', 'self'):
        for s in self_students:
            combined.append({
                'id': s.id,
                'full_name': s.full_name,
                'passport': s.passport,
                'phone': s.phone,
                'faculty': s.faculty.name if s.faculty else '-',
                'specialty': s.specialty.name if s.specialty else '-',
                'status': s.status,
                'status_display': s.get_status_display(),
                'submitted_at': s.created_at,
                'source': 'self',
                'source_display': 'O\'zi ro\'yxatdan o\'tgan',
                'added_by': '-',
                'detail_url': f'/process/{s.id}/',
            })

    # Sana bo'yicha tartiblash (yangilari yuqorida) - None qiymatlarni oxiriga
    from django.utils import timezone
    _now = timezone.now()
    combined.sort(key=lambda x: x['submitted_at'] or _now, reverse=True)

    # Statistika
    operator_count = Student.objects.exclude(created_by__isnull=True).count()
    self_count = Student.objects.filter(created_by__isnull=True).count()
    total_count = operator_count + self_count

    # Pagination
    from django.core.paginator import Paginator
    per_page = int(request.GET.get('per_page', 25))
    per_page = min(max(per_page, 10), 200)
    paginator = Paginator(combined, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'all_applicants_list.html', {
        'records': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'source_filter': source_filter,
        'status_filter': status_filter,
        'search': search,
        'operator_count': operator_count,
        'self_count': self_count,
        'total_count': total_count,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

@login_required
@axes_dispatch
def process_application(request, pk):
    active_role = request.session.get('active_role', request.user.profile.role)
    if active_role not in ['admin', 'super_admin']:
        messages.error(request, "Sizda bu bo'limni ko'rish huquqi yo'q.")
        return redirect('dashboard')

    application = get_object_or_404(Application, pk=pk)
    if request.method == 'POST':
        form = ApplicationReviewForm(request.POST, instance=application)
        if form.is_valid():
            app = form.save(commit=False)
            app.reviewed_by = request.user
            app.reviewed_at = timezone.now()
            app.save()

            # Agar Application talabaga bog'liq bo'lsa, Student status ni ham yangilash
            if app.passport:
                try:
                    student = Student.objects.filter(passport=app.passport).first()
                    if student:
                        student.status = app.status
                        student.approved_by = request.user
                        student.approved_at = timezone.now()
                        student.save()
                except Exception:
                    pass

            log_action(request.user, 'Ariza holati o\'zgartirildi', f'ID: {app.id}, Yangi holat: {app.get_status_display()}')
            messages.success(request, 'Ariza holati yangilandi.')
            return redirect('applications_list')
    else:
        form = ApplicationReviewForm(instance=application)

    return render(request, 'process_application.html', {
        'application': application,
        'form': form,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

# --- Dashboard (Axes bilan himoyalangan) ---
@login_required
@axes_dispatch
def dashboard(request):
    user = request.user
    # Session dan active_role ni olish, bo'lmasa profile.role dan foydalanish
    role = request.session.get('active_role', user.profile.role if hasattr(user, 'profile') else 'operator')

    # Talaba yoki Xodim (qara ro'yhat) arizasini ko'rish sahifasiga yo'naltirish
    if role in ['student', 'employee']:
        return redirect('my_application')

    # Faol logotipni session ga yuklash
    get_active_logo_url(request)

    log_action(user, 'Dashboard ko\'rildi')
    
    # active_role bilan filterlash
    students = get_filtered_students(user, role).prefetch_related('created_by', 'approved_by')
    
    query = request.GET.get('q')
    if query:
        students = students.filter(
            Q(full_name__icontains=query) | 
            Q(passport__icontains=query) |
            Q(jshshir__icontains=query)
        )
    
    status_filter = request.GET.get('status')
    if status_filter:
        students = students.filter(status=status_filter)

    # Advanced filtrlar
    course_filter = request.GET.get('course')
    if course_filter:
        students = students.filter(course=course_filter)

    faculty_filter = request.GET.get('faculty')
    if faculty_filter:
        students = students.filter(faculty_id=faculty_filter)

    specialty_filter = request.GET.get('specialty')
    if specialty_filter:
        students = students.filter(specialty_id=specialty_filter)

    date_from = request.GET.get('date_from')
    if date_from:
        try:
            students = students.filter(created_at__date__gte=date_from)
        except Exception:
            pass

    date_to = request.GET.get('date_to')
    if date_to:
        try:
            students = students.filter(created_at__date__lte=date_to)
        except Exception:
            pass

    source_filter = request.GET.get('source')  # 'operator' yoki 'self'
    if source_filter == 'operator':
        students = students.exclude(created_by__isnull=True)
    elif source_filter == 'self':
        students = students.filter(created_by__isnull=True)

    # Tartiblash (yangilari tepada)
    students = students.order_by('-created_at')

    all_students = get_filtered_students(user, role).prefetch_related('created_by', 'approved_by')
    stats = {
        'total': all_students.count(),
        'pending': all_students.filter(status='pending').count(),
        'approved': all_students.filter(status='approved').count(),
        'rejected': all_students.filter(status='rejected').count(),
    }

    operator_stats = None
    if role in ['super_admin', 'admin']:
        operator_stats = Student.objects.values('created_by__username').annotate(
            total=Count('id'),
            approved=Count('id', filter=Q(status='approved')),
            rejected=Count('id', filter=Q(status='rejected')),
            pending=Count('id', filter=Q(status='pending'))
        )

    recent_logs = None
    blocked_attempts = None
    if role == 'super_admin':
        recent_logs = AuditLog.objects.select_related('user')[:10]
        from axes.models import AccessAttempt
        blocked_attempts = AccessAttempt.objects.all().order_by('-attempt_time')[:5]

    unread_notifications = Notification.objects.filter(recipient=user, is_read=False).count()

    # ============ GRAFIKLAR UCHUN DATA ============
    from django.db.models.functions import TruncMonth
    from datetime import timedelta
    import json

    # Oxirgi 6 oy - ro'yxatdan o'tish statistikasi
    six_months_ago = timezone.now() - timedelta(days=180)
    monthly_data = all_students.filter(created_at__gte=six_months_ago) \
        .annotate(month=TruncMonth('created_at')) \
        .values('month') \
        .annotate(count=Count('id')) \
        .order_by('month')

    months_labels = []
    months_counts = []
    uz_months = ['Yan', 'Fev', 'Mar', 'Apr', 'May', 'Iyn', 'Iyl', 'Avg', 'Sen', 'Okt', 'Noy', 'Dek']
    for item in monthly_data:
        m = item['month']
        if m:
            months_labels.append(f"{uz_months[m.month - 1]} {m.year}")
            months_counts.append(item['count'])

    # Fakultetlar bo'yicha taqsimot (TOP 8)
    faculty_data = all_students.filter(faculty__isnull=False) \
        .values('faculty__name') \
        .annotate(count=Count('id')) \
        .order_by('-count')[:8]
    faculty_labels = [f['faculty__name'] or 'Noma\'lum' for f in faculty_data]
    faculty_counts = [f['count'] for f in faculty_data]

    # Kurs bo'yicha taqsimot
    course_data = all_students.values('course').annotate(count=Count('id')).order_by('course')
    course_labels = [f"{c['course']}-kurs" for c in course_data]
    course_counts = [c['count'] for c in course_data]

    # Status bo'yicha taqsimot (pie chart uchun)
    status_counts = [stats['pending'], stats['approved'], stats['rejected']]

    charts_data = {
        'monthly_labels': months_labels,
        'monthly_counts': months_counts,
        'faculty_labels': faculty_labels,
        'faculty_counts': faculty_counts,
        'course_labels': course_labels,
        'course_counts': course_counts,
        'status_counts': status_counts,
    }

    # Pagination - har sahifada 25 talaba
    from django.core.paginator import Paginator
    per_page = int(request.GET.get('per_page', 25))
    per_page = min(max(per_page, 10), 200)  # 10-200 oraliq
    paginator = Paginator(students, per_page)
    page_number = request.GET.get('page', 1)
    students_page = paginator.get_page(page_number)

    # Advanced filter uchun fakultet va mutaxassislik ro'yxati
    all_faculties = Faculty.objects.all().order_by('name')
    all_specialties = Specialty.objects.all().order_by('name')

    context = {
        'students': students_page,
        'page_obj': students_page,
        'paginator': paginator,
        'per_page': per_page,
        'stats': stats,
        'charts_data_json': json.dumps(charts_data),
        'all_faculties': all_faculties,
        'all_specialties': all_specialties,
        'role': role,
        'operator_stats': operator_stats,
        'recent_logs': recent_logs,
        'blocked_attempts': blocked_attempts,
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(user),
    }
    return render(request, 'dashboard.html', context)

@login_required
def notifications_view(request):
    notifications = Notification.objects.filter(recipient=request.user).order_by('-created_at')
    Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    
    context = {
        'notifications': notifications,
        'unread_notifications': 0,
        'unread_chat_count': get_unread_chat_count(request.user),
    }
    return render(request, 'notifications.html', context)

# --- Talaba Qo'shish ---
@login_required
def add_student(request):
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role in ['admin', 'hemis_id_adder', 'talaba_viewer', 'archive', 'student', 'employee']:
        log_action(request.user, 'Noqulay rol talaba qo\'shishga urindi', f'Rol: {active_role}')
        role_name = dict(request.user.profile.ROLE_CHOICES).get(active_role, 'Noma\'lum')
        messages.error(request, f"{role_name} talaba qo'sha olmaydi. Faqat operator qo'sha oladi.")
        return redirect('dashboard')

    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save(commit=False)
            student.created_by = request.user
            
            operator_comment = request.POST.get('operator_comment')
            if operator_comment:
                student.operator_comment = operator_comment
            
            student.save()
            log_action(request.user, "Talaba qo'shildi", f'Passport: {student.passport}, Ism: {student.full_name}')

            # Talaba uchun avtomatik User va Profile yaratish
            # Login va parol - pasport raqami
            try:
                # Agar bu pasport bilan user mavjud bo'lmasa, yangi yarat
                if not User.objects.filter(username=student.passport).exists():
                    new_user = User.objects.create_user(
                        username=student.passport,
                        password=student.passport,  # Parol = pasport raqami
                        first_name=student.full_name,
                        email=student.phone + '@temp.local'  # Vaqtinchalik email
                    )
                    # Profile yaratish va student rolini berish
                    profile, created = Profile.objects.get_or_create(
                        user=new_user,
                        defaults={
                            'role': 'student',
                            'roles': ['student'],
                            'jshshir': student.jshshir,
                            'phone': student.phone,
                            'email': student.phone + '@temp.local'
                        }
                    )
                    if not created:
                        profile.role = 'student'
                        profile.roles = ['student']
                        profile.jshshir = student.jshshir
                        profile.phone = student.phone
                        profile.save()

                    # Application yaratish - real status (Student.status dan olinadi)
                    # Yangi qo'shilgan talaba odatda 'pending' bo'ladi (admin tasdiqlashi kerak)
                    Application.objects.create(
                        user=new_user,
                        full_name=student.full_name,
                        passport=student.passport,
                        email=student.phone + '@temp.local',
                        phone=student.phone,
                        faculty=student.faculty,
                        specialty=student.specialty,
                        status=student.status,  # Student modelining haqiqiy statusi
                    )

                    messages.info(request, f"Talaba uchun tizimda avtomatik profil yaratildi. Login: {student.passport}, Parol: {student.passport}")
                else:
                    messages.warning(request, f"Bu pasport raqami ({student.passport}) bilan foydalanuvchi allaqachon mavjud. Profil yaratilmadi.")
            except Exception as e:
                messages.error(request, f"Talaba profili yaratishda xatolik: {str(e)}")

            # Yuklangan fayllar uchun xabar
            uploaded_files = []
            if request.FILES.get('transcript'):
                uploaded_files.append("Transkript")
            if request.FILES.get('passport_scan'):
                uploaded_files.append("Pasport nusxasi")
            if request.FILES.get('additional_documents'):
                uploaded_files.append("Qo'shimcha hujjatlar")
            
            if uploaded_files:
                messages.success(request, f"Talaba muvaffaqiyatli qo'shildi! Yuklangan fayllar: {', '.join(uploaded_files)}")
            else:
                messages.success(request, "Talaba muvaffaqiyatli qo'shildi!")
            
            return redirect('dashboard')
        else:
            log_action(request.user, "Talaba qo'shishda xatolik", f'Xatolar: {form.errors}')
            messages.error(request, "Ma'lumotlarni to'g'ri kiriting!")
    else:
        form = StudentForm()
    
    return render(request, 'add_student.html', {
        'form': form,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

# --- Talabani Tahrirlash ---
@login_required
def edit_student(request, pk):
    student = get_object_or_404(Student, pk=pk)
    
    # Session dan active_role ni olish
    role = request.session.get('active_role', request.user.profile.role if hasattr(request.user, 'profile') else 'operator')
    
    # Huquq tekshirish
    if role == 'operator':
        # Operator faqat o'zining talabalarni tahrirlay oladi
        if student.created_by != request.user:
            log_action(request.user, 'Boshqa operator talabasini tahrirlashga urindi', f'Talaba: {student.passport}')
            messages.error(request, "Siz faqat o'zingiz qo'shgan talabani tahrirlay olasiz.")
            return redirect('dashboard')
    elif role == 'admin':
        # Admin biriktirilgan operatori bo'lsa faqat o'sha operatorlarining talabalarni tahrirlay oladi
        subordinate_count = request.user.subordinate_operators.count()
        if subordinate_count > 0:
            if student.created_by and hasattr(student.created_by, 'profile'):
                if student.created_by.profile.assigned_admin != request.user:
                    log_action(request.user, 'Boshqa admin operatori talabasini tahrirlashga urindi', 
                              f'Talaba: {student.passport}')
                    messages.error(request, "Bu talabani tahrirlab bo'lmaysiz.")
                    return redirect('dashboard')
        # Aks holda (biriktirilgan operatori bo'lmaganda) barcha talabalarni tahrirlay oladi
    elif role in ['hemis_id_adder', 'talaba_viewer', 'archive']:
        log_action(request.user, 'Noqulay rol talaba tahrirlashga urindi', f'Rol: {role}')
        messages.error(request, "Siz talabalarni tahrirlay olmaysiz.")
        return redirect('dashboard')
    
    if not student.can_operator_edit():
        log_action(request.user, 'Tasdiqlangan talabani tahrirlashga urindi', 
                  f'Talaba: {student.passport}, Holat: {student.status}')
        messages.error(request, "Bu talaba admin tomonidan tasdiqlangan. Tahrirlab bo'lmaydi.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            operator_comment = request.POST.get('operator_comment')
            student = form.save(commit=False)
            student.operator_comment = operator_comment
            student.save()
            log_action(request.user, "Talaba tahrirlandi", 
                      f'Passport: {student.passport}, Yangi ma\'lumotlar saqlandi')
            
            # Yuklangan fayllar uchun xabar
            uploaded_files = []
            if request.FILES.get('transcript'):
                uploaded_files.append("Transkript")
            if request.FILES.get('passport_scan'):
                uploaded_files.append("Pasport nusxasi")
            if request.FILES.get('additional_documents'):
                uploaded_files.append("Qo'shimcha hujjatlar")
            
            if uploaded_files:
                messages.success(request, f"Talaba muvaffaqiyatli yangilandi! Yuklangan fayllar: {', '.join(uploaded_files)}")
            else:
                messages.success(request, "Talaba muvaffaqiyatli yangilandi!")
            
            return redirect('dashboard')
        else:
            messages.error(request, "Ma'lumotlarni to'g'ri kiriting!")
    else:
        form = StudentForm(instance=student)
    
    return render(request, 'edit_student.html', {
        'form': form, 
        'student': student,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

# --- EMAIL YUBORISH HELPER ---
def send_application_email(application, status, site_url=None):
    """Ariza holati o'zgarganda email yuborish"""
    from django.conf import settings

    if not application.email:
        return False

    if site_url is None:
        site_url = getattr(settings, 'SITE_URL', 'http://localhost:8000')

    subject = f"Arizangiz {'Qabul Qilindi' if status == 'approved' else 'Rad Etildi'}"

    if status == 'approved':
        template = 'emails/application_approved.html'
    else:
        template = 'emails/application_rejected.html'

    html_content = render_to_string(template, {
        'application': application,
        'site_url': site_url,
        'year': datetime.now().year,
    })

    try:
        send_mail(
            subject=subject,
            message='',  # Plain text bo'sh, HTML ishlayapti
            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@minihemis.uz'),
            recipient_list=[application.email],
            html_message=html_content,
            fail_silently=True,  # Console backend uchun muhim emas
        )
        return True
    except Exception as e:
        print(f"Email yuborish xatoligi: {e}")
        return False


# --- Tasdiqlash / Rad etish ---
@login_required
def process_student(request, pk):
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role in ['operator', 'talaba_viewer']:
        log_action(request.user, 'Noqulay rol tasdiqlash sahifasiga kirdi', 'Ruxsat yo\'q')
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')

    student = get_object_or_404(Student, pk=pk)
    
    # IDOR himoyasi: admin faqat o'z operatorlari talabalarini tasdiqlay oladi
    if active_role == 'admin':
        subordinate_count = request.user.subordinate_operators.count()
        if subordinate_count > 0:
            # Biriktirilgan operatori bor
            if student.created_by and hasattr(student.created_by, 'profile'):
                if student.created_by.profile.assigned_admin != request.user:
                    log_action(request.user, 'Boshqa admin operatori talabasini tasdiqlashga urindi', 
                              f'Talaba: {student.passport}')
                    messages.error(request, "Bu talaba sizning operatorlaringiz tomonidan qo'shilmagan.")
                    return redirect('dashboard')
        # Aks holda (biriktirilgan operatori bo'lmaganda) barcha talabalarni tasdiqlashi mumkin
    
    # IDOR himoyasi: super_admin barcha talabalarni ko'ra oladi
    if active_role not in ['super_admin', 'admin']:
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')

    if request.method == 'POST':
        old_status = student.status
        status = request.POST.get('status')
        comment = request.POST.get('comment')
        
        student.status = status
        student.comment = comment
        
        if status == 'approved':
            student.approved_at = timezone.now()
            student.approved_by = request.user
            log_action(request.user, "Talaba TASDIQLANDI",
                      f'Passport: {student.passport}, Operator: {student.created_by.username}')
            messages.success(request, f"Talaba muvaffaqiyatli TASDIQLANDI!")

            if student.created_by:
                send_notification(
                    recipient=student.created_by,
                    title="Talaba Tasdiqlandi",
                    message=f"Siz qo'shgan {student.full_name} talabasi tasdiqlandi.",
                    notification_type='student_approved',
                    sender=request.user,
                    related_object_id=student.id
                )

            # Email yuborish (agar email maydoni mavjud va qiymati bo'lsa)
            if hasattr(student, 'email') and student.email:
                send_application_email(student, 'approved')

        elif status == 'rejected':
            student.approved_by = request.user
            log_action(request.user, "Talaba RAD ETILDI",
                      f'Passport: {student.passport}, Sabab: {comment}')
            messages.warning(request, f"Talaba RAD ETILDI!")

            if student.created_by:
                send_notification(
                    recipient=student.created_by,
                    title="Talaba Rad Etildi",
                    message=f"Siz qo'shgan {student.full_name} talabasi rad etildi. Sabab: {comment}",
                    notification_type='student_rejected',
                    sender=request.user,
                    related_object_id=student.id
                )

            # Email yuborish (agar email maydoni mavjud va qiymati bo'lsa)
            if hasattr(student, 'email') and student.email:
                send_application_email(student, 'rejected')
        else:
            log_action(request.user, "Talaba holati o'zgartirildi", 
                      f'Passport: {student.passport}, {old_status} -> {status}')
            messages.info(request, f"Talaba holati o'zgartirildi.")
        
        student.save()

        # Student status o'zgarganda, shu passportga tegishli Application ham yangilansin
        try:
            applications = Application.objects.filter(passport=student.passport)
            for app in applications:
                app.status = status
                app.comment = comment
                app.reviewed_by = request.user
                app.reviewed_at = timezone.now()
                app.save()
        except Exception as e:
            log_action(request.user, "Application sinxronlashda xatolik", str(e))

        unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
        return render(request, 'process_student.html', {
            'student': student, 
            'form': CommentForm(instance=student),
            'unread_notifications': unread_notifications,
            'unread_chat_count': get_unread_chat_count(request.user),
        })

    form = CommentForm(instance=student)
    unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
    return render(request, 'process_student.html', {
        'student': student, 
        'form': form,
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

# --- O'chirish ---
@login_required
def delete_student(request, pk):
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role in ['operator', 'talaba_viewer', 'archive']:
        log_action(request.user, 'Noqulay rol talaba o\'chirishga urindi', 'Ruxsat yo\'q')
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')
    
    student = get_object_or_404(Student, pk=pk)
    
    # IDOR himoyasi: faqat super_admin va admin o'chira oladi
    if active_role not in ['super_admin', 'admin']:
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')
    
    if active_role == 'admin':
        # Admin biriktirilgan operatori bo'lsa faqat o'sha operatorlarining talabalarini o'chirishi mumkin
        subordinate_count = request.user.subordinate_operators.count()
        if subordinate_count > 0:
            if student.created_by and hasattr(student.created_by, 'profile'):
                if student.created_by.profile.assigned_admin != request.user:
                    log_action(request.user, 'Boshqa admin operatori talabasini o\'chirishga urindi', 
                              f'Talaba: {student.passport}')
                    messages.error(request, "Bu talabani o'chirish huquqingiz yo'q.")
                    return redirect('dashboard')

    passport = student.passport
    student_name = student.full_name
    student.delete()
    log_action(request.user, "Talaba O'CHIRILDI", f'Passport: {passport}, Ism: {student_name}')
    messages.success(request, "Talaba o'chirildi.")
    return redirect('dashboard')

# --- Bulk amallar (bir nechta talabani bir vaqtda boshqarish) ---
@login_required
def bulk_student_action(request):
    """Bulk action: approve/reject/delete bir nechta talabani birdan"""
    if request.method != 'POST':
        return redirect('dashboard')

    active_role = request.session.get('active_role', request.user.profile.role)
    if active_role not in ['admin', 'super_admin']:
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')

    action = request.POST.get('action')
    student_ids = request.POST.getlist('student_ids')

    if not student_ids:
        messages.warning(request, "Hech qanday talaba tanlanmadi.")
        return redirect('dashboard')

    if action not in ('approve', 'reject', 'delete'):
        messages.error(request, "Noto'g'ri amal.")
        return redirect('dashboard')

    # Admin huquqini tekshirish (biriktirilgan operatorlari talabalari)
    qs = Student.objects.filter(id__in=student_ids)
    if active_role == 'admin':
        subordinate_count = request.user.subordinate_operators.count()
        if subordinate_count > 0:
            operator_ids = list(request.user.subordinate_operators.values_list('user_id', flat=True))
            qs = qs.filter(Q(created_by_id__in=operator_ids) | Q(created_by__isnull=True))

    count = 0
    action_name = {'approve': 'tasdiqlandi', 'reject': 'rad etildi', 'delete': 'o\'chirildi'}[action]

    if action == 'delete':
        count = qs.count()
        passports = list(qs.values_list('passport', flat=True))
        qs.delete()
        log_action(request.user, f"Bulk O'CHIRISH", f"{count} ta talaba: {', '.join(passports[:10])}")
    else:
        new_status = 'approved' if action == 'approve' else 'rejected'
        for student in qs:
            student.status = new_status
            if action == 'approve':
                student.approved_at = timezone.now()
                student.approved_by = request.user
            student.save()
            # Bildirishnoma yaratish
            if student.created_by:
                send_notification(
                    recipient=student.created_by,
                    title=f"Talaba {action_name}",
                    message=f"{student.full_name} ({student.passport}) {action_name}.",
                    notification_type='info',
                    sender=request.user,
                    related_object_id=student.id,
                )
            # Email yuborish (agar email mavjud bo'lsa)
            if student.email:
                send_application_email(student, new_status)
            count += 1
        log_action(request.user, f"Bulk {action.upper()}", f"{count} ta talaba {action_name}")

    messages.success(request, f"{count} ta talaba {action_name}.")

    # Filtrlarni saqlab qaytish
    redirect_url = request.POST.get('redirect_url') or 'dashboard'
    if redirect_url.startswith('/'):
        return redirect(redirect_url)
    return redirect(redirect_url)


# --- Kommentlar ---
@login_required
def add_comment(request):
    """Talaba yoki ariza uchun komment qo'shish"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    content = request.POST.get('content')
    comment_type = request.POST.get('comment_type', 'student')
    student_id = request.POST.get('student_id')
    application_id = request.POST.get('application_id')
    is_internal = request.POST.get('is_internal', 'false') == 'true'

    if not content:
        return JsonResponse({'success': False, 'error': 'Content required'}, status=400)

    comment = Comment(
        author=request.user,
        comment_type=comment_type,
        content=content,
        is_internal=is_internal
    )

    if comment_type == 'student' and student_id:
        comment.student_id = student_id
    elif comment_type == 'application' and application_id:
        comment.application_id = application_id
    else:
        return JsonResponse({'success': False, 'error': 'Target required'}, status=400)

    comment.save()

    # Bildirishnoma yaratish (agar internal bo'lmasa va boshqa userga tegishli bo'lsa)
    if not is_internal:
        target_user = None
        if comment.student and comment.student.created_by and comment.student.created_by != request.user:
            target_user = comment.student.created_by
        elif comment.application and comment.application.user and comment.application.user != request.user:
            target_user = comment.application.user

        if target_user:
            send_notification(
                recipient=target_user,
                title="Yangi komment",
                message=f"{request.user.username} sizga komment yozdi.",
                notification_type='comment',
                sender=request.user,
                related_object_id=comment.id
            )

    return JsonResponse({
        'success': True,
        'comment': {
            'id': comment.id,
            'content': comment.content,
            'author': comment.author.username,
            'created_at': comment.created_at.strftime('%d.%m.%Y %H:%M'),
            'is_internal': comment.is_internal,
            'avatar': comment.author.profile.avatar.url if comment.author.profile.avatar else None
        }
    })


# --- HEMIS ID Qo'shish ---
@login_required
def hemis_id_list(request):
    """HEMIS ID qo'shuvchi uchun tasdiqlangan talabalar ro'yxati"""
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role not in ['hemis_id_adder', 'super_admin']:
        log_action(request.user, 'HEMIS ID ro\'yxatiga ruxsatsiz kira olmoqchi bo\'ldi', 'Ruxsat yo\'q')
        messages.error(request, "Sizda bu sahifaga kirish huquqi yo'q.")
        return redirect('dashboard')
    
    students = Student.objects.filter(status='approved').order_by('-approved_at')
    
    # Filter by search
    search_query = request.GET.get('search', '')
    if search_query:
        students = students.filter(
            Q(full_name__icontains=search_query) | 
            Q(passport__icontains=search_query) |
            Q(jshshir__icontains=search_query) |
            Q(hemis_id__icontains=search_query)
        )
    
    context = {
        'students': students,
        'search_query': search_query,
        'unread_chat_count': get_unread_chat_count(request.user),
    }
    return render(request, 'hemis_id_list.html', context)

@login_required
def update_hemis_id(request, pk):
    """Talabaning HEMIS IDsini yangilash"""
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role not in ['hemis_id_adder', 'super_admin']:
        log_action(request.user, 'HEMIS IDni o\'zgartirishga ruxsatsiz urindi', 'Ruxsat yo\'q')
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')
    
    student = get_object_or_404(Student, pk=pk)
    
    if student.status != 'approved':
        messages.error(request, "Faqat tasdiqlangan talabalarning HEMIS IDsini qo'shish mumkin.")
        return redirect('hemis_id_list')
    
    if request.method == 'POST':
        hemis_id = request.POST.get('hemis_id', '').strip()
        
        if not hemis_id:
            messages.error(request, "HEMIS ID bo'sh bo'lsa bo'lmaydi!")
            return redirect('hemis_id_list')
        
        if len(hemis_id) > 20:
            messages.error(request, "HEMIS ID 20 ta belgidan ko'p bo'lsa bo'lmaydi!")
            return redirect('hemis_id_list')
        
        # Agar bu HEMIS ID boshqa talabada borsa
        existing = Student.objects.filter(hemis_id=hemis_id).exclude(pk=pk).first()
        if existing:
            messages.error(request, f"Bu HEMIS ID allaqachon {existing.full_name} uchun ishlatilgan!")
            return redirect('hemis_id_list')
        
        student.hemis_id = hemis_id
        student.save()
        
        log_action(request.user, "HEMIS ID qo'shildi", 
                  f'Talaba: {student.full_name}, HEMIS ID: {hemis_id}')
        messages.success(request, f"HEMIS ID muvaffaqiyatli qo'shildi: {hemis_id}")
        return redirect('hemis_id_list')
    
    context = {
        'student': student,
        'unread_chat_count': get_unread_chat_count(request.user),
    }
    return render(request, 'update_hemis_id.html', context)

# --- Excel Yuklash ---
@login_required
def export_excel(request):
    # Session dan active_role ni olish
    role = request.session.get('active_role', request.user.profile.role)
    
    if role in ['admin', 'super_admin']:
        log_action(request.user, "Excel yuklandi (Admin)")
        return _generate_excel(request)
    
    if role == 'operator':
        from django.db import transaction
        
        # Race condition himoyasi: transaction va select_for_update
        with transaction.atomic():
            approved_request = ExcelDownloadRequest.objects.select_for_update().filter(
                requested_by=request.user,
                status='approved'
            ).first()
            
            if not approved_request:
                log_action(request.user, "Excel yuklashga urindi (ruxsat yo'q)")
                messages.error(request, "Excel yuklash uchun admin ruxsati kerak! Iltimos, so'rov yuboring.")
                return redirect('request_excel')
            
            approved_request.download_count += 1
            approved_request.save()
        
        log_action(request.user, "Excel yuklandi (Operator)", 
                  f'So\'rov ID: {approved_request.id}, Yuklash: {approved_request.download_count}')
        messages.success(request, "Excel yuklanmoqda...")
        
        return _generate_excel(request)
    
    messages.error(request, "Ruxsat yo'q")
    return redirect('dashboard')

def _generate_excel(request):
    role = request.session.get('active_role', request.user.profile.role if hasattr(request.user, 'profile') else 'operator')
    students = get_filtered_students(request.user, role)

    query = request.GET.get('q')
    if query:
        students = students.filter(
            Q(full_name__icontains=query) |
            Q(passport__icontains=query)
        )

    status_filter = request.GET.get('status')
    if status_filter:
        students = students.filter(status=status_filter)

    # Advanced filtrlar
    if request.GET.get('course'):
        students = students.filter(course=request.GET.get('course'))
    if request.GET.get('faculty'):
        students = students.filter(faculty_id=request.GET.get('faculty'))
    if request.GET.get('specialty'):
        students = students.filter(specialty_id=request.GET.get('specialty'))
    if request.GET.get('date_from'):
        try:
            students = students.filter(created_at__date__gte=request.GET.get('date_from'))
        except Exception:
            pass
    if request.GET.get('date_to'):
        try:
            students = students.filter(created_at__date__lte=request.GET.get('date_to'))
        except Exception:
            pass
    src = request.GET.get('source')
    if src == 'operator':
        students = students.exclude(created_by__isnull=True)
    elif src == 'self':
        students = students.filter(created_by__isnull=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'talabalar_{datetime.now().strftime("%Y-%m-%d_%H-%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Talabalar"
    
    headers = [
        "ID", "Ism Familiya", "Pasport", "JSHSHIR", "Kurs",
        "Telefon", "Fakultet", "Mutaxassislik", "Holat",
        "Manba", "Qo'shgan", "Sana", "Operator Izohi", "Admin Izohi", "Tasdiqlangan Sana"
    ]
    ws.append(headers)

    # Qo'shgan odam ism familiyasini olish helper
    def _added_by_name(user):
        if not user:
            return "-"
        full = f"{user.first_name} {user.last_name}".strip()
        return full if full else user.username

    # Operator/admin tomonidan qo'shilgan talabalar
    for s in students:
        ws.append([
            s.id,
            s.full_name,
            s.passport,
            s.jshshir,
            s.course,
            s.phone,
            s.faculty.name if s.faculty else "-",
            s.specialty.name if s.specialty else "-",
            s.get_status_display(),
            "Operator qo'shgan",
            _added_by_name(s.created_by),
            s.created_at.strftime('%Y-%m-%d %H:%M'),
            s.operator_comment if s.operator_comment else "-",
            s.comment if s.comment else "-",
            s.approved_at.strftime('%Y-%m-%d %H:%M') if s.approved_at else "-"
        ])

    # Admin/Super admin uchun: o'zi ro'yxatdan o'tgan talabalarni ham qo'shish
    if role in ['admin', 'super_admin']:
        operator_passports = list(Student.objects.values_list('passport', flat=True))
        self_apps = Application.objects.select_related('user', 'faculty', 'specialty') \
            .exclude(passport__in=operator_passports) \
            .filter(faculty__isnull=False, specialty__isnull=False)  # faqat talabalar (xodim emas)

        if status_filter:
            self_apps = self_apps.filter(status=status_filter)
        if query:
            self_apps = self_apps.filter(
                Q(full_name__icontains=query) | Q(passport__icontains=query)
            )

        for app in self_apps:
            ws.append([
                f"A{app.id}",  # Application ID prefix
                app.full_name,
                app.passport or "-",
                "-",  # JSHSHIR
                "-",  # Kurs
                app.phone,
                app.faculty.name if app.faculty else "-",
                app.specialty.name if app.specialty else "-",
                app.get_status_display(),
                "O'zi ro'yxatdan o'tgan",
                "-",  # Qo'shgan - o'zi
                app.submitted_at.strftime('%Y-%m-%d %H:%M'),
                "-",
                app.comment if app.comment else "-",
                app.reviewed_at.strftime('%Y-%m-%d %H:%M') if app.reviewed_at else "-"
            ])
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # BytesIO-ga saqlash
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    
    return response

def _generate_sample_excel():
    """Generate sample Excel file with only required fields for upload"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'talabalar_namuna_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Namuna"
    
    config = ExcelImportConfig.get_active_config()
    from openpyxl.utils import column_index_from_string
    full_name_idx = column_index_from_string(config.full_name_col)
    passport_idx = column_index_from_string(config.passport_col)
    jshshir_idx = column_index_from_string(config.jshshir_col)
    course_idx = column_index_from_string(config.course_col)
    phone_idx = column_index_from_string(config.phone_col)
    faculty_idx = column_index_from_string(config.faculty_col)
    specialty_idx = column_index_from_string(config.specialty_col)
    operator_comment_idx = column_index_from_string(config.operator_comment_col)
    
    # Write headers
    ws.cell(row=1, column=full_name_idx, value="Ism Familiya")
    ws.cell(row=1, column=passport_idx, value="Pasport")
    ws.cell(row=1, column=jshshir_idx, value="JSHSHIR")
    ws.cell(row=1, column=course_idx, value="Kurs")
    ws.cell(row=1, column=phone_idx, value="Telefon")
    ws.cell(row=1, column=faculty_idx, value="Fakultet")
    ws.cell(row=1, column=specialty_idx, value="Mutaxassislik")
    ws.cell(row=1, column=operator_comment_idx, value="Operator izohi")
    
    # Style the header
    for cell in ws[1]:
        if cell.value:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50) if max_length > 0 else 10
        ws.column_dimensions[column].width = adjusted_width
    
    # BytesIO-ga saqlash
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    
    return response

def _generate_yoriqnoma_excel():
    """Generate yo'riqnoma Excel file with columns according to layout"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'yoriqnoma_namuna_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yo'riqnoma"
    
    config = ExcelImportConfig.get_active_config()
    from openpyxl.utils import column_index_from_string
    full_name_idx = column_index_from_string(config.full_name_col)
    passport_idx = column_index_from_string(config.passport_col)
    jshshir_idx = column_index_from_string(config.jshshir_col)
    course_idx = column_index_from_string(config.course_col)
    phone_idx = column_index_from_string(config.phone_col)
    faculty_idx = column_index_from_string(config.faculty_col)
    specialty_idx = column_index_from_string(config.specialty_col)
    operator_comment_idx = column_index_from_string(config.operator_comment_col)
    
    # Yo'riqnoma formatdagi ustunlar
    yoriqnoma_headers = {
        'full_name': "To'liq ism familiya",
        'passport': "Pasport seriya va raqami (AA1234567)",
        'jshshir': "JSHSHIR (14 ta raqam)",
        'course': "Kurs (1, 2, 3, 4)",
        'phone': "Telefon raqam",
        'faculty': "Fakultet nomi",
        'specialty': "Mutaxassislik nomi",
        'operator_comment': "Operator izohi (ixtiyoriy)"
    }
    
    ws.cell(row=1, column=full_name_idx, value=yoriqnoma_headers['full_name'])
    ws.cell(row=1, column=passport_idx, value=yoriqnoma_headers['passport'])
    ws.cell(row=1, column=jshshir_idx, value=yoriqnoma_headers['jshshir'])
    ws.cell(row=1, column=course_idx, value=yoriqnoma_headers['course'])
    ws.cell(row=1, column=phone_idx, value=yoriqnoma_headers['phone'])
    ws.cell(row=1, column=faculty_idx, value=yoriqnoma_headers['faculty'])
    ws.cell(row=1, column=specialty_idx, value=yoriqnoma_headers['specialty'])
    ws.cell(row=1, column=operator_comment_idx, value=yoriqnoma_headers['operator_comment'])
    
    # Namuna ma'lumotlar
    sample_row_1 = {
        'full_name': 'Aliyev Ali Vali o\'g\'li',
        'passport': 'AA1234567',
        'jshshir': '12345678901234',
        'course': '3',
        'phone': '998901234567',
        'faculty': 'Informatika va axborot texnologiyalari fakulteti',
        'specialty': 'Kompyuter injiniring',
        'operator_comment': 'A\'lo talaba'
    }
    sample_row_2 = {
        'full_name': 'Karimova Karima O\'tkir qizi',
        'passport': 'BB2345678',
        'jshshir': '23456789012345',
        'course': '2',
        'phone': '998902345678',
        'faculty': 'Matematika fakulteti',
        'specialty': 'Amaliy matematika',
        'operator_comment': 'Yaxshi natijalar'
    }
    
    # Write sample 1 (row 2)
    ws.cell(row=2, column=full_name_idx, value=sample_row_1['full_name'])
    ws.cell(row=2, column=passport_idx, value=sample_row_1['passport'])
    ws.cell(row=2, column=jshshir_idx, value=sample_row_1['jshshir'])
    ws.cell(row=2, column=course_idx, value=sample_row_1['course'])
    ws.cell(row=2, column=phone_idx, value=sample_row_1['phone'])
    ws.cell(row=2, column=faculty_idx, value=sample_row_1['faculty'])
    ws.cell(row=2, column=specialty_idx, value=sample_row_1['specialty'])
    ws.cell(row=2, column=operator_comment_idx, value=sample_row_1['operator_comment'])
    
    # Write sample 2 (row 3)
    ws.cell(row=3, column=full_name_idx, value=sample_row_2['full_name'])
    ws.cell(row=3, column=passport_idx, value=sample_row_2['passport'])
    ws.cell(row=3, column=jshshir_idx, value=sample_row_2['jshshir'])
    ws.cell(row=3, column=course_idx, value=sample_row_2['course'])
    ws.cell(row=3, column=phone_idx, value=sample_row_2['phone'])
    ws.cell(row=3, column=faculty_idx, value=sample_row_2['faculty'])
    ws.cell(row=3, column=specialty_idx, value=sample_row_2['specialty'])
    ws.cell(row=3, column=operator_comment_idx, value=sample_row_2['operator_comment'])
    
    # Style the header
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for cell in ws[1]:
        if cell.value:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    column_widths = {
        full_name_idx: 30,
        passport_idx: 25,
        jshshir_idx: 20,
        course_idx: 15,
        phone_idx: 18,
        faculty_idx: 35,
        specialty_idx: 30,
        operator_comment_idx: 25
    }
    
    for col_idx, width in column_widths.items():
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
    
    # Style data rows
    for row in ws.iter_rows():
        for cell in row:
            if cell.row > 1:  # Not header
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.font = Font(size=11)
    
    # BytesIO-ga saqlash
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    
    return response

def _generate_failed_students_excel(failed_students, duplicate_passports):
    """Generate Excel file for failed and duplicate students with error details"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'xatolikli_talabalar_{datetime.now().strftime("%Y-%m-%d_%H-%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb = openpyxl.Workbook()
    
    # Failed students sheet
    ws_failed = wb.active
    ws_failed.title = "Xatoliklar"
    
    config = ExcelImportConfig.get_active_config()
    from openpyxl.utils import column_index_from_string
    
    # Qator column is always at Column A (1)
    ws_failed.cell(row=1, column=1, value="Qator")
    
    full_name_idx = column_index_from_string(config.full_name_col) + 1
    passport_idx = column_index_from_string(config.passport_col) + 1
    jshshir_idx = column_index_from_string(config.jshshir_col) + 1
    course_idx = column_index_from_string(config.course_col) + 1
    phone_idx = column_index_from_string(config.phone_col) + 1
    faculty_idx = column_index_from_string(config.faculty_col) + 1
    specialty_idx = column_index_from_string(config.specialty_col) + 1
    operator_comment_idx = column_index_from_string(config.operator_comment_col) + 1
    
    ws_failed.cell(row=1, column=full_name_idx, value="Ism Familiya")
    ws_failed.cell(row=1, column=passport_idx, value="Pasport")
    ws_failed.cell(row=1, column=jshshir_idx, value="JSHSHIR")
    ws_failed.cell(row=1, column=course_idx, value="Kurs")
    ws_failed.cell(row=1, column=phone_idx, value="Telefon")
    ws_failed.cell(row=1, column=faculty_idx, value="Fakultet")
    ws_failed.cell(row=1, column=specialty_idx, value="Mutaxassislik")
    ws_failed.cell(row=1, column=operator_comment_idx, value="Operator izohi")
    
    # The error column is placed after the last column
    max_idx = max(full_name_idx, passport_idx, jshshir_idx, course_idx, phone_idx, faculty_idx, specialty_idx, operator_comment_idx)
    error_col_idx = max_idx + 1
    ws_failed.cell(row=1, column=error_col_idx, value="Xatolik sababi")
    
    # Write data rows
    def get_val(row, idx):
        if 0 <= idx < len(row):
            return row[idx]
        return ""
        
    orig_full_name_idx = column_index_from_string(config.full_name_col) - 1
    orig_passport_idx = column_index_from_string(config.passport_col) - 1
    orig_jshshir_idx = column_index_from_string(config.jshshir_col) - 1
    orig_course_idx = column_index_from_string(config.course_col) - 1
    orig_phone_idx = column_index_from_string(config.phone_col) - 1
    orig_faculty_idx = column_index_from_string(config.faculty_col) - 1
    orig_specialty_idx = column_index_from_string(config.specialty_col) - 1
    orig_operator_comment_idx = column_index_from_string(config.operator_comment_col) - 1

    for r_idx, student in enumerate(failed_students, start=2):
        data = student['data']
        ws_failed.cell(row=r_idx, column=1, value=student['row'])
        ws_failed.cell(row=r_idx, column=full_name_idx, value=get_val(data, orig_full_name_idx))
        ws_failed.cell(row=r_idx, column=passport_idx, value=get_val(data, orig_passport_idx))
        ws_failed.cell(row=r_idx, column=jshshir_idx, value=get_val(data, orig_jshshir_idx))
        ws_failed.cell(row=r_idx, column=course_idx, value=get_val(data, orig_course_idx))
        ws_failed.cell(row=r_idx, column=phone_idx, value=get_val(data, orig_phone_idx))
        ws_failed.cell(row=r_idx, column=faculty_idx, value=get_val(data, orig_faculty_idx))
        ws_failed.cell(row=r_idx, column=specialty_idx, value=get_val(data, orig_specialty_idx))
        ws_failed.cell(row=r_idx, column=operator_comment_idx, value=get_val(data, orig_operator_comment_idx))
        ws_failed.cell(row=r_idx, column=error_col_idx, value=student['reason'])
    
    # Duplicate students sheet
    if duplicate_passports:
        ws_duplicate = wb.create_sheet("Dublikatlar")
        headers_duplicate = [
            "Qator", "Ism Familiya", "Pasport", "Xatolik sababi"
        ]
        ws_duplicate.append(headers_duplicate)
        
        for student in duplicate_passports:
            ws_duplicate.append([
                student['row'],
                student['full_name'],
                student['passport'],
                "Pasport seriyasi bazada mavjud"
            ])
    
    # Style headers
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            if cell.value:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # Auto-adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50) if max_length > 0 else 10
            sheet.column_dimensions[column].width = adjusted_width
    
    # BytesIO-ga saqlash
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    
    return response

# --- Namuna Excel Yuklash (Admin ruxsatisiz) ---
@login_required
def download_sample_excel(request):
    """Namuna Excel faylini yuklab olish (admin ruxsatisiz)"""
    log_action(request.user, "Namuna Excel yuklandi")
    return _generate_sample_excel()

# --- Yo'riqnoma Excel Yuklash ---
def download_yoriqnoma_excel(request):
    """Yo'riqnoma Excel faylini yuklab olish"""
    return _generate_yoriqnoma_excel()

# --- Xatolikli Talabalarni Excelga Yuklash ---
@login_required
def download_failed_students_excel(request):
    """Xatolikli talabalarni Excel faylga yuklab olish"""
    # Sessiondan ma'lumotlarni olish
    failed_students = request.session.get('failed_students', [])
    duplicate_passports = request.session.get('duplicate_passports', [])
    
    if not failed_students and not duplicate_passports:
        messages.error(request, "Yuklash uchun xatolikli talabalar mavjud emas!")
        return redirect('upload_students_excel')
    
    log_action(request.user, "Xatolikli talabalar Excel yuklandi")
    return _generate_failed_students_excel(failed_students, duplicate_passports)

# --- Excel Yuklash So'rovi ---
@login_required
def request_excel(request):
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role == 'operator':
        if request.method == 'POST':
            form = ExcelRequestForm(request.POST)
            if form.is_valid():
                excel_request = form.save(commit=False)
                excel_request.requested_by = request.user
                excel_request.save()
                
                log_action(request.user, "Excel yuklash so'rovi yuborildi", 
                          f'Sabab: {excel_request.reason}')
                
                admins = User.objects.filter(profile__role__in=['admin', 'super_admin'])
                
                xabarlar_soni = 0
                for admin in admins:
                    send_notification(
                        recipient=admin,
                        title="🔔 Yangi Excel Yuklash So'rovi",
                        message=f"{request.user.username} operator Excel yuklash uchun so'rov yubordi.\n\nSabab: {excel_request.reason}",
                        notification_type='excel_request',
                        sender=request.user,
                        related_object_id=excel_request.id
                    )
                    xabarlar_soni += 1
                
                messages.success(request, f"So'rov yuborildi! {xabarlar_soni} ta adminga xabar bordi.")
                return redirect('excel_requests_list')
        else:
            form = ExcelRequestForm()
        
        return render(request, 'request_excel.html', {
            'form': form,
            'unread_chat_count': get_unread_chat_count(request.user),
        })
    else:
        messages.error(request, "Faqat operatorlar Excel yuklash so'rovini yuborishi mumkin.")
        return redirect('dashboard')

# --- Excel So'rovlarini Ko'rish ---
@login_required
def excel_requests_list(request):
    # Session dan active_role ni olish
    role = request.session.get('active_role', request.user.profile.role)
    
    if role == 'operator':
        excel_requests = ExcelDownloadRequest.objects.filter(requested_by=request.user)
    elif role in ['admin', 'super_admin']:
        if role == 'admin':
            operator_ids = request.user.subordinate_operators.values_list('user_id', flat=True)
            excel_requests = ExcelDownloadRequest.objects.filter(requested_by__id__in=operator_ids)
        else:
            excel_requests = ExcelDownloadRequest.objects.all()
    else:
        messages.error(request, "Ruxsat yo'q")
        return redirect('dashboard')
    
    status_filter = request.GET.get('status')
    if status_filter:
        excel_requests = excel_requests.filter(status=status_filter)
    
    unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
    
    context = {
        'excel_requests': excel_requests,
        'role': role,
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(request.user),
    }
    return render(request, 'excel_requests_list.html', context)

# --- Excel So'rovini Tasdiqlash ---
@login_required
def process_excel_request(request, pk):
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role == 'operator':
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')
    
    excel_request = get_object_or_404(ExcelDownloadRequest, pk=pk)
    
    if active_role == 'admin':
        if excel_request.requested_by.profile.assigned_admin != request.user:
            messages.error(request, "Bu so'rov sizning operatoringizdan emas.")
            return redirect('excel_requests_list')
    
    if request.method == 'POST':
        status = request.POST.get('status')
        admin_comment = request.POST.get('admin_comment')
        
        excel_request.status = status
        excel_request.admin_comment = admin_comment
        
        if status == 'approved':
            excel_request.approved_at = timezone.now()
            excel_request.approved_by = request.user
            log_action(request.user, "Excel yuklash TASDIQLANDI", 
                      f'Operator: {excel_request.requested_by.username}')
            messages.success(request, "So'rov TASDIQLANDI!")
            
            send_notification(
                recipient=excel_request.requested_by,
                title="✅ Excel So'rovi Tasdiqlandi",
                message=f"Sizning Excel yuklash so'rovingiz tasdiqlandi. Endi yuklashingiz mumkin.",
                notification_type='excel_approved',
                sender=request.user,
                related_object_id=excel_request.id
            )
        else:
            log_action(request.user, "Excel yuklash RAD ETILDI", 
                      f'Operator: {excel_request.requested_by.username}, Sabab: {admin_comment}')
            messages.warning(request, "So'rov RAD ETILDI!")
            
            send_notification(
                recipient=excel_request.requested_by,
                title="❌ Excel So'rovi Rad Etildi",
                message=f"Sizning Excel yuklash so'rovingiz rad etildi. Sabab: {admin_comment}",
                notification_type='excel_rejected',
                sender=request.user,
                related_object_id=excel_request.id
            )
        
        excel_request.save()
        return redirect('excel_requests_list')
    
    form = ExcelAdminForm(instance=excel_request)
    unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
    return render(request, 'process_excel_request.html', {
        'excel_request': excel_request, 
        'form': form,
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

# --- Profil Sozlamalari ---
@login_required
def profile_settings(request):
    password_form = PasswordChangeForm()
    username_form = UsernameChangeForm(instance=request.user, user=request.user)
    profile_form = ProfileEditForm(instance=request.user, profile=request.user.profile)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'change_password':
            password_form = PasswordChangeForm(request.POST)
            if password_form.is_valid():
                if request.user.check_password(password_form.cleaned_data['old_password']):
                    request.user.set_password(password_form.cleaned_data['new_password'])
                    request.user.save()
                    update_session_auth_hash(request, request.user)
                    log_action(request.user, "Parol o'zgartirildi")
                    messages.success(request, "Parol muvaffaqiyatli o'zgartirildi!")
                    return redirect('dashboard')
                else:
                    log_action(request.user, "Parol o'zgartirishda xatolik", "Eski parol noto'g'ri")
                    messages.error(request, "Eski parol noto'g'ri")
            else:
                messages.error(request, "Yangi parol talablarga javob bermaydi")
                
        elif action == 'change_username':
            username_form = UsernameChangeForm(request.POST, instance=request.user, user=request.user)
            if username_form.is_valid():
                old_username = request.user.username
                username_form.save()
                log_action(request.user, "Username o'zgartirildi", f"{old_username} -> {request.user.username}")
                messages.success(request, f"Login nomi '{request.user.username}' ga muvaffaqiyatli o'zgartirildi!")
                return redirect('dashboard')
                
        elif action == 'edit_profile':
            profile_form = ProfileEditForm(request.POST, request.FILES, instance=request.user, profile=request.user.profile)
            if profile_form.is_valid():
                # Eski emailni saqlash
                old_email = request.user.profile.email
                
                # Profile ni saqlash
                profile_form.save()
                
                # Email o'zgarganligini tekshirish
                new_email = request.user.profile.email
                
                if old_email != new_email:
                    # Email o'zgargan bo'lsa, log qilish va xabar berish
                    log_action(request.user, "Email o'zgartirildi", f"{old_email} -> {new_email}")
                    messages.success(request, f"Profil ma'lumotlaringiz muvaffaqiyatli o'zgartirildi! Email '{new_email}' ga yangilandi.")
                    
                    # Cache ni tozalash
                    from django.core.cache import cache
                    cache.clear()
                    
                    print(f"DEBUG: Email o'zgardi: {old_email} -> {new_email}")
                else:
                    log_action(request.user, "Profil ma'lumotlari o'zgartirildi")
                    messages.success(request, "Profil ma'lumotlaringiz muvaffaqiyatli o'zgartirildi!")
                
                return redirect('profile_settings')
    
    unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
    return render(request, 'profile.html', {
        'password_form': password_form,
        'username_form': username_form,
        'profile_form': profile_form,
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

# --- Rolni Almashtirish (Switch Role) ---
@login_required
def switch_role(request, role):
    """Foydalanuvchi rolini almashtirish"""
    # Tekshirish - foydalanuvchida bu rol borligini
    if not request.user.profile.has_role(role):
        messages.error(request, "Sizda bu rol yo'q!")
        return redirect('dashboard')

    # Rolni session ga saqlash
    request.session['active_role'] = role
    request.session.modified = True  # Session o'zgarganini majburlash

    role_display = dict(Profile.ROLE_CHOICES).get(role, role)
    messages.success(request, f"Rol o'zgartirildi: {role_display}")
    return redirect('dashboard')

# --- Foydalanuvchi Boshqaruvi (Super Admin uchun) ---
@login_required
def manage_users(request):
    """Foydalanuvchilar ro'yxatini ko'rish va boshqarish (Super Admin)"""
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)

    if active_role != 'super_admin':
        messages.error(request, "Sizda bu huquq yo'q!")
        return redirect('dashboard')

    # Foydalanuvchi turini olish (student, employee, staff, all)
    user_type = request.GET.get('type', 'all')  # 'all', 'student', 'employee', 'staff'

    users = User.objects.filter(profile__isnull=False).select_related('profile').order_by('-date_joined')

    # Foydalanuvchi turiga qarab filtrlash
    # Student: role='student'
    # Employee: role='employee'
    # Staff: boshqa rollar (operator, admin, etc.)
    if user_type == 'student':
        users = users.filter(profile__role='student')
    elif user_type == 'employee':
        users = users.filter(profile__role='employee')
    elif user_type == 'staff':
        users = users.exclude(profile__role__in=['student', 'employee'])

    # Filter by role (only check primary role field - SQLite doesn't support JSONField contains)
    role_filter = request.GET.get('role')
    if role_filter:
        users = users.filter(profile__role=role_filter)

    # Search
    search = request.GET.get('search')
    if search:
        users = users.filter(
            Q(username__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search)
        )

    # Statistika
    total_count = users.count()
    student_count = User.objects.filter(profile__role='student').count()
    employee_count = User.objects.filter(profile__role='employee').count()
    staff_count = User.objects.exclude(profile__role__in=['student', 'employee']).filter(profile__isnull=False).count()

    return render(request, 'manage_users.html', {
        'users': users,
        'role_filter': role_filter,
        'search': search,
        'user_type': user_type,
        'total_count': total_count,
        'student_count': student_count,
        'employee_count': employee_count,
        'staff_count': staff_count,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

@login_required
def create_user(request):
    """Yangi foydalanuvchi yaratish (Super Admin)"""
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role != 'super_admin':
        messages.error(request, "Sizda bu huquq yo'q!")
        return redirect('dashboard')
    
    if request.method == 'POST':
        print(f"DEBUG: POST data received: {request.POST}")
        form = UserRegistrationForm(request.POST)
        print(f"DEBUG: Form created, errors before validation: {form.errors}")
        if form.is_valid():
            print(f"DEBUG: Form is valid, saving...")
            try:
                user = form.save()

                # Qo'shimcha ma'lumotlarni saqlash (ixtiyoriy)
                user.first_name = request.POST.get('first_name', '')
                user.last_name = request.POST.get('last_name', '')
                user.email = request.POST.get('email', '')
                user.save()

                # Profile ma'lumotlari
                if hasattr(user, 'profile'):
                    user.profile.phone = request.POST.get('phone', '')
                    user.profile.save()

                log_action(request.user, "Foydalanuvchi yaratildi", f"Username: {user.username}, Rol: {user.profile.role}")
                messages.success(request, f"Foydalanuvchi '{user.username}' muvaffaqiyatli yaratildi!")
                return redirect('manage_users')
            except Exception as e:
                messages.error(request, f"Xatolik yuz berdi: {str(e)}")
                print(f"DEBUG create_user error: {e}")
        else:
            print(f"DEBUG: Form is NOT valid, errors: {form.errors}")
            # Form xatoliklarini ko'rsatish
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f"Xatolik: {error}")
                    else:
                        messages.error(request, f"{field}: {error}")
    else:
        form = UserRegistrationForm()
    
    return render(request, 'create_user.html', {
        'form': form,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

@login_required
def edit_user(request, user_id):
    """Foydalanuvchi ma'lumotlarini tahrirlash (Super Admin)"""
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role != 'super_admin':
        messages.error(request, "Sizda bu huquq yo'q!")
        return redirect('dashboard')
    
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_info':
            user.first_name = request.POST.get('first_name', '')
            user.last_name = request.POST.get('last_name', '')
            user.email = request.POST.get('email', '')
            user.save()

            profile = user.profile
            profile.phone = request.POST.get('phone', '')

            # Handle organizational_role (tashkiliy rol)
            organizational_role = request.POST.get('organizational_role')
            profile.organizational_role = organizational_role if organizational_role else None

            # Handle multiple roles
            selected_roles = request.POST.getlist('roles')
            if selected_roles:
                profile.roles = selected_roles
                # Set the first role as the primary role for backward compatibility
                profile.role = selected_roles[0]
            else:
                # Fallback to single role if no multiple roles selected
                single_role = request.POST.get('role', profile.role)
                profile.role = single_role
                profile.roles = [single_role]

            profile.save()

            org_role_display = profile.get_organizational_role_display() if profile.organizational_role else "Yo'q"
            log_action(request.user, "Foydalanuvchi tahrirlandi", f"Username: {user.username}, Roles: {profile.roles}, Tashkiliy Rol: {org_role_display}")
            messages.success(request, "Foydalanuvchi ma'lumotlari yangilandi!")
            
        elif action == 'reset_password':
            new_password = request.POST.get('new_password')
            if new_password:
                user.set_password(new_password)
                user.save()
                log_action(request.user, "Foydalanuvchi paroli tiklandi", f"Username: {user.username}")
                messages.success(request, f"'{user.username}' paroli muvaffaqiyatli o'zgartirildi!")
        
        elif action == 'delete_user':
            username = user.username
            user.delete()
            log_action(request.user, "Foydalanuvchi o'chirildi", f"Username: {username}")
            messages.success(request, f"Foydalanuvchi '{username}' o'chirildi!")
            return redirect('manage_users')
        
        return redirect('edit_user', user_id=user_id)
    
    return render(request, 'edit_user.html', {
        'edit_user': user,
        'unread_chat_count': get_unread_chat_count(request.user),
        'organizational_role_choices': Profile.ORGANIZATIONAL_ROLE_CHOICES,
    })

# --- Fakultet Boshqaruvi ---
@login_required
def manage_faculties(request):
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role == 'operator':
        log_action(request.user, 'Operator fakultet boshqaruviga kirdi', 'Ruxsat yo\'q')
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        faculty_name = request.POST.get('faculty_name')
        specialty_name = request.POST.get('specialty_name')
        faculty_id = request.POST.get('faculty_id')
        
        if faculty_name:
            Faculty.objects.create(name=faculty_name)
            log_action(request.user, "Fakultet qo'shildi", faculty_name)
            messages.success(request, "Fakultet qo'shildi")
        
        if specialty_name and faculty_id:
            faculty = get_object_or_404(Faculty, id=faculty_id)
            contract_amount = request.POST.get('contract_amount', 0)
            try:
                contract_amount = int(contract_amount)
            except (ValueError, TypeError):
                contract_amount = 0
            Specialty.objects.create(faculty=faculty, name=specialty_name, contract_amount=contract_amount)
            log_action(request.user, "Mutaxassislik qo'shildi", f'{faculty.name} - {specialty_name} ({contract_amount:,} so\'m)')
            messages.success(request, f"Mutaxassislik qo'shildi! Kontrakt: {contract_amount:,} so'm")
    
    faculties = Faculty.objects.all()
    specialties = Specialty.objects.all()
    
    unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
    
    return render(request, 'manage_faculties.html', {
        'faculties': faculties,
        'specialties': specialties,
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

# --- Fakultetni Tahrirlash ---
@login_required
def edit_faculty(request, pk):
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role == 'operator':
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')
    
    faculty = get_object_or_404(Faculty, pk=pk)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        
        if name:
            old_name = faculty.name
            faculty.name = name
            faculty.save()
            
            log_action(request.user, "Fakultet tahrirlandi", 
                      f'{old_name} -> {name}')
            messages.success(request, "Fakultet yangilandi!")
            return redirect('manage_faculties')
        else:
            messages.error(request, "Fakultet nomini kiriting!")
    
    unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
    return render(request, 'edit_faculty.html', {
        'faculty': faculty,
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

# --- Fakultetni O'chirish ---
@login_required
def delete_faculty(request, pk):
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role == 'operator':
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')
    
    faculty = get_object_or_404(Faculty, pk=pk)
    faculty_name = faculty.name
    
    specialties_count = faculty.specialty_set.count()
    if specialties_count > 0:
        messages.error(request, f"Bu fakultetda {specialties_count} ta mutaxassislik bor. Avval ularni o'chiring!")
        return redirect('manage_faculties')
    
    faculty.delete()
    log_action(request.user, "Fakultet o'chirildi", faculty_name)
    messages.success(request, f"'{faculty_name}' fakulteti o'chirildi!")
    return redirect('manage_faculties')

# --- Mutaxassislikni Tahrirlash ---
@login_required
def edit_specialty(request, pk):
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role == 'operator':
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')
    
    specialty = get_object_or_404(Specialty, pk=pk)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        faculty_id = request.POST.get('faculty')
        
        if name and faculty_id:
            specialty.name = name
            specialty.faculty_id = faculty_id
            specialty.save()
            
            log_action(request.user, "Mutaxassislik tahrirlandi", 
                      f'{specialty.name} (ID: {pk})')
            messages.success(request, "Mutaxassislik yangilandi!")
            return redirect('manage_faculties')
        else:
            messages.error(request, "Barcha maydonlarni to'ldiring!")
    
    faculties = Faculty.objects.all()
    unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
    return render(request, 'edit_specialty.html', {
        'specialty': specialty,
        'faculties': faculties,
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

# --- Mutaxassislikni O'chirish ---
@login_required
def delete_specialty(request, pk):
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role == 'operator':
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')
    
    specialty = get_object_or_404(Specialty, pk=pk)
    specialty_name = specialty.name
    
    students_count = Student.objects.filter(specialty=specialty).count()
    if students_count > 0:
        messages.error(request, f"Bu mutaxassislikka {students_count} ta talaba biriktirilgan. O'chirib bo'lmaydi!")
        return redirect('manage_faculties')
    
    specialty.delete()
    log_action(request.user, "Mutaxassislik o'chirildi", specialty_name)
    messages.success(request, f"'{specialty_name}' mutaxassisligi o'chirildi!")
    return redirect('manage_faculties')

# --- Mutaxassislik Kontrakt Summasini Yangilash ---
@login_required
def update_specialty_contract(request, pk):
    """Mutaxassislik kontrakt summasini yangilash (Super Admin)"""
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role != 'super_admin':
        messages.error(request, "Sizda bu huquq yo'q! Faqat Super Admin kontrakt summasini o'zgartirishi mumkin.")
        return redirect('manage_faculties')
    
    specialty = get_object_or_404(Specialty, pk=pk)
    
    if request.method == 'POST':
        contract_amount = request.POST.get('contract_amount')
        try:
            contract_amount = int(contract_amount)
            if contract_amount < 0:
                messages.error(request, "Kontrakt summasi manfiy bo'lishi mumkin emas!")
                return redirect('manage_faculties')
            
            old_amount = specialty.contract_amount
            specialty.contract_amount = contract_amount
            specialty.save()
            
            log_action(request.user, "Mutaxassislik kontrakt summasi yangilandi", 
                      f'{specialty.name}: {old_amount} -> {contract_amount}')
            messages.success(request, f"'{specialty.name}' mutaxassisligi kontrakt summasi {contract_amount:,} so'mga yangilandi!")
        except (ValueError, TypeError):
            messages.error(request, "Noto'g'ri summa kiritildi!")
    
    return redirect('manage_faculties')

# --- Audit Loglar ---
@login_required
def audit_logs_view(request):
    if not is_super_admin(request.user):
        log_action(request.user, 'Super admin bo\'lmagan audit loglarga kirdi', 'Ruxsat yo\'q')
        messages.error(request, "Faqat Super Admin ko'rishi mumkin.")
        return redirect('dashboard')
    
    logs = AuditLog.objects.select_related('user').all()
    
    user_filter = request.GET.get('user')
    action_filter = request.GET.get('action')
    
    # Qidirish maydonlarini kengaytirish - pasport, ism familiya
    search_query = request.GET.get('search', '')
    if search_query:
        logs = logs.filter(
            Q(user__username__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(action__icontains=search_query) |
            Q(details__icontains=search_query)
        )
    
    if user_filter:
        logs = logs.filter(user__username__icontains=user_filter)
    if action_filter:
        logs = logs.filter(action__icontains=action_filter)
    
    # Yuklab olish funksiyasi
    if request.GET.get('download') == 'excel':
        return download_audit_logs_excel(logs, search_query)
    
    logs = logs[:200]
    
    unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
    
    return render(request, 'audit_logs.html', {
        'logs': logs,
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(request.user),
        'search_query': search_query,
    })

def download_audit_logs_excel(logs, search_query=''):
    """Audit loglarni Excel fayliga yuklab olish"""
    import openpyxl
    from django.http import HttpResponse
    from datetime import datetime
    
    # Excel fayl yaratish
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Loglar"
    
    # Sarlavhalar
    headers = ['ID', 'Vaqt', 'Foydalanuvchi', 'Amal', 'Tafsilotlar']
    ws.append(headers)
    
    # Ma'lumotlarni yozish
    for log in logs:
        row = [
            log.id,
            log.timestamp.strftime('%d.%m.%Y %H:%M:%S'),
            log.user.username if log.user else 'System',
            log.action,
            log.details or ''
        ]
        ws.append(row)
    
    # Styling
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
    
    # Column width
    column_widths = [8, 20, 15, 30, 50]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"audit_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    if search_query:
        filename += f"_search_{search_query}"
    filename += ".xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response

# --- Bloklangan Foydalanuvchilar ---
@login_required
def blocked_users_view(request):
    if not is_super_admin(request.user):
        log_action(request.user, 'Super admin bo\'lmagan bloklangan foydalanuvchilarga kirdi', 'Ruxsat yo\'q')
        messages.error(request, "Faqat Super Admin ko'rishi mumkin.")
        return redirect('dashboard')
    
    from axes.models import AccessAttempt
    
    # Bloklangan foydalanuvchilar
    blocked_attempts = AccessAttempt.objects.all().order_by('-attempt_time')
    
    # Qidirish
    search_query = request.GET.get('search', '')
    if search_query:
        blocked_attempts = blocked_attempts.filter(
            Q(username__icontains=search_query) |
            Q(ip_address__icontains=search_query)
        )
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(blocked_attempts, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Yuklab olish
    if request.GET.get('download') == 'excel':
        return download_blocked_users_excel(blocked_attempts, search_query)
    
    unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
    
    return render(request, 'blocked_users.html', {
        'page_obj': page_obj,
        'search_query': search_query,
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

def download_blocked_users_excel(attempts, search_query=''):
    """Bloklangan foydalanuvchilarni Excel fayliga yuklab olish"""
    import openpyxl
    from django.http import HttpResponse
    from datetime import datetime
    
    # Excel fayl yaratish
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bloklangan Foydalanuvchilar"
    
    # Sarlavhalar
    headers = ['Username', 'IP Address', 'Urinishlar Soni', 'Vaqt', 'User Agent', 'Fail Reason']
    ws.append(headers)
    
    # Ma'lumotlarni yozish
    for attempt in attempts:
        row = [
            attempt.username or '',
            attempt.ip_address or '',
            attempt.failures_since_start or 0,
            attempt.attempt_time.strftime('%d.%m.%Y %H:%M:%S') if attempt.attempt_time else '',
            attempt.user_agent or '',
            attempt.failure_reason or ''
        ]
        ws.append(row)
    
    # Styling
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
        cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
    
    # Column width
    column_widths = [15, 15, 10, 20, 30, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"blocked_users_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    if search_query:
        filename += f"_search_{search_query}"
    filename += ".xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response

@login_required
def unlock_user_direct(request, username):
    """Foydalanuvchini blokdan chiqarish (to'g'ridan-to'g'ri)"""
    if not is_super_admin(request.user):
        messages.error(request, "Faqat Super Admin blokdan chiqarishi mumkin.")
        return redirect('dashboard')
    
    try:
        from axes.models import AccessAttempt
        from django.contrib.auth.models import User
        from django.core.cache import cache
        from django.db.models import Q
        
        # AccessAttempt larni o'chirish (username yoki ip_address bo'yicha)
        deleted_count, _ = AccessAttempt.objects.filter(Q(username=username) | Q(ip_address=username)).delete()
        
        # Cache dan ham IP blokirovkasini va rate limitlarni o'chirish
        cache.delete(f'blocked_ip:{username}')
        cache.delete(f'rate_limit:general:{username}:anon')
        cache.delete(f'rate_limit:login:{username}:anon')
        cache.delete(f'rate_limit:api:{username}:anon')
        
        # Ro'yxatdan o'tgan foydalanuvchilar bo'lsa ularning limitlarini ham tozalaymiz
        if User.objects.filter(username=username).exists():
            cache.delete(f'rate_limit:general:{username}:{username}')
            cache.delete(f'rate_limit:login:{username}:{username}')
            cache.delete(f'rate_limit:api:{username}:{username}')
        
        messages.success(request, f"✅ {username} blokdan chiqarildi va uning kesh xotirasi tozalandi!")
        log_action(request.user, f'Foydalanuvchi/IP blokdan chiqarildi', f'{username} blokdan chiqarildi va kesh tozalandi')
        
    except Exception as e:
        messages.error(request, f"❌ Xatolik: {str(e)}")
    
    return redirect('blocked_users')

@login_required
def unblock_super_admin_profile(request):
    """Super admin profilini blokdan chiqarish"""
    if not is_super_admin(request.user):
        messages.error(request, "Faqat Super Admin foydalanuvchi profilini blokdan chiqarishi mumkin.")
        return redirect('dashboard')
    
    try:
        from axes.models import AccessAttempt
        from django.contrib.auth.models import User
        
        # O'zining super admin profilini blokdan chiqarish
        deleted_count, _ = AccessAttempt.objects.filter(username=request.user.username).delete()
        
        if deleted_count > 0:
            messages.success(request, f"✅ Sizning profilingiz blokdan chiqarildi! ({deleted_count} ta urinish o'chirildi)")
            log_action(request.user, 'Super admin o\'z profilini blokdan chiqardi', f'{deleted_count} ta urinish o\'chirildi')
        else:
            messages.info(request, "ℹ️ Sizning profilingiz bloklangan emas")
        
    except Exception as e:
        messages.error(request, f"❌ Xatolik: {str(e)}")
    
    return redirect('dashboard')

@login_required
def clear_all_super_admin_blocks(request):
    """Barcha super admin larni blokdan chiqarish"""
    if not is_super_admin(request.user):
        messages.error(request, "Faqat Super Admin barcha super admin larni blokdan chiqarishi mumkin.")
        return redirect('dashboard')
    
    try:
        from axes.models import AccessAttempt
        from django.contrib.auth.models import User
        
        # Barcha super admin larni topish
        super_admins = User.objects.filter(profile__role='super_admin')
        total_deleted = 0
        
        for admin in super_admins:
            deleted_count, _ = AccessAttempt.objects.filter(username=admin.username).delete()
            total_deleted += deleted_count
        
        if total_deleted > 0:
            messages.success(request, f"✅ Barcha Super Admin lar blokdan chiqarildi! ({total_deleted} ta urinish o'chirildi)")
            log_action(request.user, 'Barcha super admin larni blokdan chiqardi', f'{total_deleted} ta urinish o\'chirildi')
        else:
            messages.info(request, "ℹ️ Bloklangan Super Admin lar yo'q")
        
    except Exception as e:
        messages.error(request, f"❌ Xatolik: {str(e)}")
    
    return redirect('dashboard')

@login_required
def clear_all_blocks(request):
    """Barcha bloklangan foydalanuvchilar va API larni ochirish"""
    if not is_super_admin(request.user):
        messages.error(request, "Faqat Super Admin barcha bloklarni o'chirishi mumkin.")
        return redirect('dashboard')
    
    try:
        from axes.models import AccessAttempt
        
        # Barcha bloklarni o'chirish
        deleted_count, _ = AccessAttempt.objects.all().delete()
        
        if deleted_count > 0:
            messages.success(request, f"✅ Barcha bloklar o'chirildi! ({deleted_count} ta urinish o'chirildi)")
            log_action(request.user, 'Barcha bloklarni o\'chirildi', f'{deleted_count} ta urinish o\'chirildi')
        else:
            messages.info(request, "ℹ️ Bloklangan foydalanuvchilar yo'q")
        
    except Exception as e:
        messages.error(request, f"❌ Xatolik: {str(e)}")
    
    return redirect('dashboard')

# --- Email Sozlamalari (Super Admin) ---
@login_required
def email_settings_view(request):
    """Email sozlamalari paneli"""
    if not is_super_admin(request.user):
        messages.error(request, "Faqat Super Admin email sozlamalarini o'zgartira oladi.")
        return redirect('dashboard')
    
    from .models import EmailSettings
    
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        smtp_host = request.POST.get('smtp_host', 'smtp.gmail.com')
        smtp_port = request.POST.get('smtp_port', 587)
        use_tls = request.POST.get('use_tls') == 'on'
        
        if not email or not password:
            messages.error(request, 'Email manzili va parolni kiriting!')
        else:
            # Eski sozlamalarni nofaol qilish
            EmailSettings.objects.all().update(is_active=False)
            
            # Yangi sozlamalarni saqlash
            settings_obj = EmailSettings.objects.create(
                email=email,
                password=password,
                smtp_host=smtp_host,
                smtp_port=int(smtp_port),
                use_tls=use_tls,
                is_active=True
            )
            
            # Email sozlamalarini tekshirish
            try:
                send_test_email(email, password, smtp_host, smtp_port, use_tls)
                messages.success(request, 'Email sozlamalari muvaffaqiyatli saqlandi va tekshirildi!')
                log_action(request.user, 'Email sozlamalari yangilandi', f'Email: {email}')
            except Exception as e:
                messages.error(request, f'Email sozlamalari saqlandi, lekin tekshirishda xatolik: {str(e)}')
                log_action(request.user, 'Email sozlamalari yangilandi (xatolik bilan)', f'Email: {email}, Xatolik: {str(e)}')
    
    # Joriy sozlamalarni olish
    current_settings = EmailSettings.get_active_settings()
    
    return render(request, 'email_settings.html', {
        'current_settings': current_settings,
        'all_settings': EmailSettings.objects.all().order_by('-created_at')
    })


@login_required
def excel_import_settings_view(request):
    """Excel import ustunlari sozlamalari (faqat super admin uchun)"""
    if not is_super_admin(request.user):
        messages.error(request, "Faqat Super Admin bu sahifaga kira oladi.")
        return redirect('dashboard')
    
    config = ExcelImportConfig.get_active_config()
    
    if request.method == 'POST':
        form = ExcelImportConfigForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            log_action(request.user, "Excel import sozlamalari yangilandi")
            messages.success(request, "Excel import sozlamalari muvaffaqiyatli yangilandi!")
            return redirect('excel_import_settings')
        else:
            messages.error(request, "Sozlamalarni saqlashda xatolik yuz berdi. Iltimos, xatolarni to'g'rilang.")
    else:
        form = ExcelImportConfigForm(instance=config)
        
    return render(request, 'excel_import_settings.html', {
        'form': form,
        'config': config,
        'unread_chat_count': get_unread_chat_count(request.user),
    })


@login_required
def test_email_settings(request, settings_id):
    """Email sozlamalarini tekshirish"""
    if not is_super_admin(request.user):
        messages.error(request, "Faqat Super Admin email sozlamalarini tekshira oladi.")
        return redirect('dashboard')
    
    from .models import EmailSettings
    
    try:
        email_settings = EmailSettings.objects.get(id=settings_id)
        
        import os
        es_password = getattr(email_settings, 'password', None) or os.environ.get('EMAIL_HOST_PASSWORD') or ''
        
        send_test_email(
            email_settings.email,
            es_password,
            email_settings.smtp_host,
            email_settings.smtp_port,
            email_settings.use_tls
        )
        
        messages.success(request, f'{email_settings.email} uchun test email yuborildi!')
        
    except EmailSettings.DoesNotExist:
        messages.error(request, 'Email sozlamalari topilmadi!')
    except Exception as e:
        messages.error(request, f'Test email yuborishda xatolik: {str(e)}')
    
    return redirect('email_settings')

@login_required
def delete_email_settings(request, settings_id):
    """Email sozlamasini o'chirish"""
    if not is_super_admin(request.user):
        messages.error(request, "Faqat Super Admin email sozlamalarini o'chira oladi.")
        return redirect('dashboard')

    from .models import EmailSettings
    try:
        es = EmailSettings.objects.get(id=settings_id)
        if es.is_active:
            messages.error(request, f"Faol sozlama o'chirilmaydi! Avval boshqasini faollashtiring.")
        else:
            email = es.email
            es.delete()
            log_action(request.user, "Email sozlamasi o'chirildi", f'Email: {email}')
            messages.success(request, f"{email} sozlamasi o'chirildi.")
    except EmailSettings.DoesNotExist:
        messages.error(request, 'Email sozlamalari topilmadi!')
    return redirect('email_settings')


@login_required
def edit_email_settings(request, settings_id):
    """Email sozlamasini tahrirlash"""
    if not is_super_admin(request.user):
        messages.error(request, "Faqat Super Admin email sozlamalarini tahrirlay oladi.")
        return redirect('dashboard')

    from .models import EmailSettings
    try:
        es = EmailSettings.objects.get(id=settings_id)
    except EmailSettings.DoesNotExist:
        messages.error(request, 'Email sozlamalari topilmadi!')
        return redirect('email_settings')

    if request.method == 'POST':
        es.email = request.POST.get('email', es.email).strip()
        new_password = request.POST.get('password', '').strip()
        if new_password:
            es.password = new_password
        es.smtp_host = request.POST.get('smtp_host', es.smtp_host) or 'smtp.gmail.com'
        es.smtp_port = int(request.POST.get('smtp_port') or 587)
        es.use_tls = request.POST.get('use_tls') == 'on'
        es.save()
        log_action(request.user, "Email sozlamasi tahrirlandi", f'Email: {es.email}')
        messages.success(request, "Email sozlamasi yangilandi!")
        return redirect('email_settings')

    return render(request, 'email_settings_edit.html', {'setting': es})


@login_required
def toggle_email_settings(request, settings_id):
    """Email sozlamalarini faol/nofaol qilish"""
    if not is_super_admin(request.user):
        messages.error(request, "Faqat Super Admin email sozlamalarini o'zgartira oladi.")
        return redirect('dashboard')
    
    from .models import EmailSettings
    
    try:
        email_settings = EmailSettings.objects.get(id=settings_id)
        
        # Barcha sozlamalarni nofaol qilish
        EmailSettings.objects.all().update(is_active=False)
        
        # Tanlangan sozlamalarni faol qilish
        email_settings.is_active = True
        email_settings.save()
        
        messages.success(request, f'{email_settings.email} faol sozlama sifatida belgilandi!')
        log_action(request.user, 'Email sozlamalari faollashtirildi', f'Email: {email_settings.email}')
        
    except EmailSettings.DoesNotExist:
        messages.error(request, 'Email sozlamalari topilmadi!')
    
    return redirect('email_settings')

def send_test_email(email, password, smtp_host, smtp_port, use_tls):
    """Test email yuborish"""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    try:
        # SMTP serverga ulanish
        server = smtplib.SMTP(smtp_host, smtp_port)
        server.ehlo()  # Serverga salomlash
        
        if use_tls:
            server.starttls()
            server.ehlo()  # TLS dan keyin qayta salomlash
        
        # Login qilish - Unicode belgilarini olib tashlash
        clean_email = email.replace('\u200b', '').strip()
        
        # Get password from settings or environment if empty/None
        raw_password = password
        if not raw_password:
            import os
            from django.conf import settings
            raw_password = os.environ.get('EMAIL_HOST_PASSWORD') or getattr(settings, 'EMAIL_HOST_PASSWORD', '')
            
        clean_password = (raw_password or '').replace('\u200b', '').strip()
        server.login(clean_email, clean_password)
        
        # Test email yaratish
        msg = MIMEMultipart()
        msg['From'] = email
        msg['To'] = email
        msg['Subject'] = 'Test-sinov-ucun - Email Sozlamalari Testi'
        
        body = '''
        Assalomu alaykum!
        
        Bu Test-sinov-ucun tizimidan email sozlamalari testi.
        
        Agar siz bu xabarni qabul qilsangiz, email sozlamalari to'g'ri ishlaydi.
        
        Hurmat bilan,
        Test-sinov-ucun jamoasi
        '''.replace('\u200b', '').strip()
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Email yuborish
        server.send_message(msg)
        server.quit()
        
    except Exception as e:
        raise Exception(f"Email yuborish xatoliki: {str(e)}")

# --- Parolni Tiklash Tizimi ---
def forgot_password_view(request):
    """Parolni unutgan sahifasi - email kiritish"""
    from django.contrib.auth.models import User
    from django.core.cache import cache
    
    # Rate limiting: IP bo'yicha 10 daqiqada 5 ta so'rov
    ip_address = request.META.get('REMOTE_ADDR', '')
    cache_key = f'password_reset_rate_limit:{ip_address}'
    attempts = cache.get(cache_key, 0)
    
    if attempts >= 5:
        messages.error(request, 'Juda ko\'p urinish! Iltimos, 10 daqiqadan so\'ng qayta urinib ko\'ring.')
        return render(request, 'forgot_password.html')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        
        if not username:
            messages.error(request, 'Iltimos, loginni kiriting!')
            return render(request, 'forgot_password.html')
        
        # Rate limitni oshirish
        cache.set(cache_key, attempts + 1, 600)  # 10 daqiqa
        
        try:
            user = User.objects.get(username=username)
            
            # Email tekshirish - shaxsiy email (real vaqtda tekshirish)
            if not hasattr(user, 'profile'):
                messages.error(request, 'Sizning profilingiz mavjud emas. Admin bilan bog\'laning!')
                return render(request, 'forgot_password.html')
            
            # Profile ni real vaqtda tekshirish - bazadagi ma'lumotlarni to'liq tekshirish
            try:
                # Django cache ni tozalash
                from django.core.cache import cache
                cache.clear()
                
                # Bazadagi barcha ma'lumotlarni tekshirish
                from django.db import connection
                with connection.cursor() as cursor:
                    # User ID ni olish
                    cursor.execute("SELECT id FROM auth_user WHERE username = %s", [username])
                    user_result = cursor.fetchone()
                    user_id = user_result[0] if user_result else None
                    
                    print(f"DEBUG: User ID: {user_id}")
                    
                    if not user_id:
                        messages.error(request, 'Bunday login topilmadi!')
                        return render(request, 'forgot_password.html')
                    
                    # Profile ma'lumotlarini olish
                    cursor.execute("""
                        SELECT id, email, user_id FROM core_profile 
                        WHERE user_id = %s
                        LIMIT 1
                    """, [user_id])
                    profile_result = cursor.fetchone()
                    
                    print(f"DEBUG: Profile result: {profile_result}")
                    
                    if not profile_result:
                        messages.error(request, 'Profile topilmadi! Admin bilan bog\'laning!')
                        return render(request, 'forgot_password.html')
                    
                    profile_id, profile_email, profile_user_id = profile_result
                    user_email = profile_email
                    
                    if not user_email or user_email.strip() == '':
                        messages.error(request, 'Login uchun email biriktirilmagan. Iltimos, adminga murojaat qiling!')
                        return render(request, 'forgot_password.html')
                
                # Debug uchun batafsil ma'lumot
                print(f"DEBUG: User ID: {user_id}")
                print(f"DEBUG: Profile ID: {profile_id}")
                print(f"DEBUG: Profile email: {user_email}")
                print(f"DEBUG: Profile user_id: {profile_user_id}")
                
                # User obyektini olish va profile ni yangilash
                user = User.objects.get(username=username)
                user.profile.refresh_from_db()
                
                # Ikki marta tekshirish - SQL va ORM
                orm_email = user.profile.email
                print(f"DEBUG: ORM email: {orm_email}")
                print(f"DEBUG: SQL email: {user_email}")
                
                # Agar ORM email SQL dan farq qilsa, bazani yangilash
                if orm_email and orm_email != user_email:
                    print(f"DEBUG: Email farqi bor, bazani yangilash: {user_email} -> {orm_email}")
                    
                    # To'g'ridan-to'g'ri bazani yangilash
                    cursor.execute("""
                        UPDATE core_profile 
                        SET email = %s 
                        WHERE user_id = %s
                    """, [orm_email, user_id])
                    
                    print(f"DEBUG: Baza yangilandi: {orm_email}")
                    user_email = orm_email
                    print(f"DEBUG: Yangi email ishlatilmoqda: {user_email}")
                else:
                    print(f"DEBUG: SQL email ishlatilmoqda: {user_email}")
                
                # Session ga ham saqlash
                request.session['current_email'] = user_email
                
            except Exception as e:
                messages.error(request, 'Profile ma\'lumotlarini olishda xatolik. Admin bilan bog\'laning!')
                print(f"DEBUG: Profile refresh xatoligi: {str(e)}")
                import traceback
                traceback.print_exc()
                return render(request, 'forgot_password.html')
            
            # Kod generatsiya qilish
            from .models import PasswordResetCode
            reset_code = PasswordResetCode.generate_code(user)
            
            # Emailga kod yuborish - super admin emailidan
            try:
                # Debug: User topilganini tekshirish
                print(f"DEBUG: User topildi: {username}")
                
                # Super admin emailini topish
                from django.contrib.auth.models import User
                super_admin_user = User.objects.filter(profile__role='super_admin').first()
                
                if super_admin_user and hasattr(super_admin_user, 'profile') and super_admin_user.profile.email:
                    sender_email = super_admin_user.profile.email
                    print(f"DEBUG: Super admin email: {sender_email}")
                else:
                    # Agar super admin emaili bo'lmasa, default email
                    sender_email = 'noreply@minihemis.uz'
                    print(f"DEBUG: Default email ishlatilmoqda: {sender_email}")
                
                # Foydalanuvchi emailiga kod yuborish - oldin tekshirilgan emailni ishlatish
                user_email = user_email  # Yuqoridan tekshirilgan email
                print(f"DEBUG: Foydalanuvchi email (real vaqtda tekshirilgan): {user_email}")
                print(f"DEBUG: Generatsiya qilingan kod: {reset_code.code}")
                print(f"DEBUG: User ID: {user.id}, Profile ID: {user.profile.id if hasattr(user, 'profile') else 'None'}")
                
                send_password_reset_code(user_email, reset_code.code, sender_email)
                print(f"DEBUG: Email yuborish muvaffaqiyatli tugadi")
                
                # Session ga email manzilini saqlash
                request.session['reset_email'] = user_email
                request.session['reset_code'] = reset_code.code
                print(f"DEBUG: Session ga saqlandi: {user_email}, {reset_code.code}")
                
                # JavaScript uchun JSON response
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'redirect': f'/verify-reset-code/{username}/',
                        'message': f'{user_email} manziliga kod yuborildi! Emailni tekshiring.'
                    })
                else:
                    messages.success(request, f'{user_email} manziliga kod yuborildi! Emailni tekshiring.')
                    return redirect('verify_reset_code', username=username)
                
            except Exception as e:
                print(f"DEBUG: Email yuborishda xatolik: {str(e)}")
                print(f"DEBUG: Xatolik turi: {type(e)}")
                import traceback
                traceback.print_exc()
                
                # JavaScript uchun JSON response
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'error': str(e)
                    })
                else:
                    messages.error(request, f'Email yuborishda xatolik: {str(e)}')
                    return render(request, 'forgot_password.html')
                
        except User.DoesNotExist:
            messages.error(request, 'Bunday login topilmadi!')
    
    return render(request, 'forgot_password.html')

def verify_reset_code_view(request, username):
    """Kodni tasdiqlash sahifasi"""
    from django.core.cache import cache
    
    # Rate limiting: IP bo'yicha 10 daqiqada 10 ta urinish
    ip_address = request.META.get('REMOTE_ADDR', '')
    cache_key = f'verify_code_rate_limit:{ip_address}'
    attempts = cache.get(cache_key, 0)
    
    if attempts >= 10:
        messages.error(request, 'Juda ko\'p urinish! Iltimos, 10 daqiqadan so\'ng qayta urinib ko\'ring.')
        return render(request, 'verify_reset_code.html', {'username': username})
    
    if request.method == 'POST':
        code = request.POST.get('code')
        
        # Rate limitni oshirish
        cache.set(cache_key, attempts + 1, 600)  # 10 daqiqa
        
        if not code or len(code) not in [6, 8]:
            messages.error(request, 'Iltimos, 6 yoki 8 xonalik kodni kiriting!')
            return render(request, 'verify_reset_code.html', {'username': username})
        
        try:
            user = User.objects.get(username=username)
            from .models import PasswordResetCode
            
            # Kodni tekshirish
            reset_code = PasswordResetCode.objects.filter(
                user=user, 
                code=code, 
                is_used=False
            ).first()
            
            if not reset_code or not reset_code.is_valid():
                messages.error(request, 'Noto\'g\'ri yoki muddati o\'tgan kod!')
                return render(request, 'verify_reset_code.html', {'username': username})
            
            # Kodni ishlatilgan deb belgilash
            reset_code.is_used = True
            reset_code.save()
            
            messages.success(request, 'Kod tasdiqlandi! Yangi parol o\'rnating.')
            return redirect('reset_password', username=username)
            
        except User.DoesNotExist:
            messages.error(request, 'Bunday login topilmadi!')
            return redirect('forgot_password')
    
    # Session dan email manzilini olish
    reset_email = request.session.get('reset_email', '')
    reset_code = request.session.get('reset_code', '')
    
    return render(request, 'verify_reset_code.html', {
        'username': username,
        'reset_email': reset_email,
        'reset_code': reset_code
    })

def reset_password_view(request, username):
    """Yangi parol o'rnatish sahifasi"""
    if request.method == 'POST':
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        
        if not password or not confirm_password:
            messages.error(request, 'Iltimos, barcha maydonlarni to\'ldiring!')
            return render(request, 'reset_password.html', {'username': username})
        
        if password != confirm_password:
            messages.error(request, 'Parollar mos kelmadi!')
            return render(request, 'reset_password.html', {'username': username})
        
        if len(password) < 8:
            messages.error(request, 'Parol kamida 8 ta belgidan iborat bo\'lishi kerak!')
            return render(request, 'reset_password.html', {'username': username})
        
        try:
            user = User.objects.get(username=username)
            user.set_password(password)
            user.save()
            
            messages.success(request, 'Parol muvaffaqiyatli o\'zgartirildi! Endi login qiling.')
            return redirect('login')
            
        except User.DoesNotExist:
            messages.error(request, 'Bunday login topilmadi!')
            return redirect('forgot_password')
    
    return render(request, 'reset_password.html', {'username': username})

def send_password_reset_code(email, code, sender_email=None):
    """Emailga 6 xonalik kod yuborish - EmailSettings dan foydalanish"""
    from .models import EmailSettings
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    # Faol email sozlamalarini olish
    email_settings = EmailSettings.get_active_settings()
    
    if not email_settings:
        raise Exception("Faol email sozlama topilmadi! Super admin bilan bog'laning.")
    
    try:
        # SMTP serverga ulanish
        server = smtplib.SMTP(email_settings.smtp_host, email_settings.smtp_port)
        server.ehlo()  # Serverga salomlash
        
        if email_settings.use_tls:
            server.starttls()
            server.ehlo()  # TLS dan keyin qayta salomlash
        
        # Login qilish - Unicode belgilarini olib tashlash
        clean_email = email_settings.email.replace('\u200b', '').strip()
        
        # Get password from settings or environment if not present in email_settings
        raw_password = getattr(email_settings, 'password', None)
        if not raw_password:
            import os
            from django.conf import settings
            raw_password = os.environ.get('EMAIL_HOST_PASSWORD') or getattr(settings, 'EMAIL_HOST_PASSWORD', '')
            
        clean_password = (raw_password or '').replace('\u200b', '').strip()
        server.login(clean_email, clean_password)
        
        # Email yaratish
        msg = MIMEMultipart()
        msg['From'] = email_settings.email
        msg['To'] = email
        msg['Subject'] = 'Test-sinov-ucun - Parolni Tiklash Kodi'
        
        body = f'''
        Assalomu alaykum!
        
        Sizning parolni tiklash kodingiz: {code}
        
        Bu kod 15 daqiqa davomida yaroqli.
        
        Agar siz parolni tiklashni so'ramagan bo'lsangiz, ushbu xabarni e'tiborsiz qoldiring.
        
        Hurmat bilan 
        '''.replace('\u200b', '').strip()
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Email yuborish
        server.send_message(msg)
        server.quit()
        
        print(f"Email muvaffaqiyatli yuborildi: {email_settings.email} -> {email}")
        
    except Exception as e:
        print(f"Email yuborish xatoliki: {e}")
        raise e

# --- Fakultet bo'yicha Mutaxassisliklar (AJAX) ---
def get_specialties(request):
    """Fakultet bo'yicha yo'nalishlarni olish - Ro'yxatdan o'tish uchun ochiq"""
    faculty_id = request.GET.get('faculty_id')

    if faculty_id:
        specialties = Specialty.objects.filter(faculty_id=faculty_id)
        data = list(specialties.values('id', 'name'))
        return JsonResponse(data, safe=False)
    else:
        return JsonResponse([], safe=False)

# --- Operatorlar Statistikasi ---
@login_required
def operator_statistics(request):
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role not in ['admin', 'super_admin']:
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if active_role == 'super_admin':
        operators = User.objects.filter(profile__role='operator')
    else:
        operator_ids = request.user.subordinate_operators.values_list('user_id', flat=True)
        operators = User.objects.filter(id__in=operator_ids, profile__role='operator')
    
    operator_stats = []
    for operator in operators:
        students = Student.objects.filter(created_by=operator)
        
        if start_date:
            students = students.filter(created_at__date__gte=start_date)
        if end_date:
            students = students.filter(created_at__date__lte=end_date)
        
        stats = {
            'operator': operator,
            'total': students.count(),
            'pending': students.filter(status='pending').count(),
            'approved': students.filter(status='approved').count(),
            'rejected': students.filter(status='rejected').count(),
        }
        operator_stats.append(stats)
    
    context = {
        'operator_stats': operator_stats,
        'start_date': start_date,
        'end_date': end_date,
        'unread_notifications': Notification.objects.filter(recipient=request.user, is_read=False).count(),
        'unread_chat_count': get_unread_chat_count(request.user),
    }
    return render(request, 'operator_statistics.html', context)

# --- Operator Statistikasini Excel Yuklash ---
@login_required
def export_operator_statistics_excel(request):
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role not in ['admin', 'super_admin']:
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if active_role == 'super_admin':
        operators = User.objects.filter(profile__role='operator')
    else:
        operator_ids = request.user.subordinate_operators.values_list('user_id', flat=True)
        operators = User.objects.filter(id__in=operator_ids, profile__role='operator')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    if start_date and end_date:
        filename = f'operator_statistika_{start_date}_to_{end_date}.xlsx'
    else:
        filename = f'operator_statistika_barchasi.xlsx'
    
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Operatorlar Statistikasi"
    
    headers = [
        "#", "Operator", "Jami", "Qabul Qilingan", "Rad Etilgan", 
        "Kutilmoqda", "Foiz (Qabul)", "Foiz (Rad)"
    ]
    ws.append(headers)
    
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_color = Font(color="FFFFFF", bold=True, size=12)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.font = header_font_color
    
    total_all = 0
    total_approved = 0
    total_rejected = 0
    total_pending = 0
    
    for idx, operator in enumerate(operators, 1):
        students = Student.objects.filter(created_by=operator)
        
        if start_date:
            students = students.filter(created_at__date__gte=start_date)
        if end_date:
            students = students.filter(created_at__date__lte=end_date)
        
        total = students.count()
        approved = students.filter(status='approved').count()
        rejected = students.filter(status='rejected').count()
        pending = students.filter(status='pending').count()
        
        approved_percent = (approved / total * 100) if total > 0 else 0
        rejected_percent = (rejected / total * 100) if total > 0 else 0
        
        ws.append([
            idx,
            operator.username,
            total,
            approved,
            rejected,
            pending,
            f"{approved_percent:.1f}%",
            f"{rejected_percent:.1f}%"
        ])
        
        total_all += total
        total_approved += approved
        total_rejected += rejected
        total_pending += pending
    
    ws.append([])
    ws.append([
        "JAMI",
        "",
        total_all,
        total_approved,
        total_rejected,
        total_pending,
        "",
        ""
    ])
    
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width
    
    # BytesIO-ga saqlash
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    
    return response
@login_required
def export_all_students_excel(request):
    role = request.session.get('active_role', request.user.profile.role if hasattr(request.user, 'profile') else 'operator')
    if role not in ['admin', 'super_admin', 'operator']:
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    students = get_filtered_students(request.user, role)
    
    if start_date:
        students = students.filter(created_at__date__gte=start_date)
    if end_date:
        students = students.filter(created_at__date__lte=end_date)
    
    query = request.GET.get('q')
    if query:
        students = students.filter(
            Q(full_name__icontains=query) | 
            Q(passport__icontains=query)
        )
    
    status_filter = request.GET.get('status')
    if status_filter:
        students = students.filter(status=status_filter)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    if start_date and end_date:
        filename = f'talabalar_{start_date}_to_{end_date}.xlsx'
    else:
        filename = f'talabalar_barchasi_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Talabalar"
    
    headers = [
        "ID", "Ism Familiya", "Pasport", "JSHSHIR", "Kurs", 
        "Telefon", "Fakultet", "Mutaxassislik", "Holat", 
        "Qo'shgan", "Qo'shilgan Sana", "Tasdiqlangan Sana", 
        "Operator Izohi", "Admin Izohi"
    ]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for s in students:
        ws.append([
            s.id,
            s.full_name,
            s.passport,
            s.jshshir,
            s.course,
            s.phone,
            s.faculty.name if s.faculty else "-",
            s.specialty.name if s.specialty else "-",
            s.get_status_display(),
            s.created_by.username if s.created_by else "-",
            s.created_at.strftime('%Y-%m-%d %H:%M'),
            s.approved_at.strftime('%Y-%m-%d %H:%M') if s.approved_at else "-",
            s.operator_comment if s.operator_comment else "-",
            s.comment if s.comment else "-"
        ])
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column].width = adjusted_width
    
    # BytesIO-ga saqlash
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    
    return response

# --- Eng Yaxshi Operatorlar (Ranking) ---
@login_required
def operator_ranking(request):
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role not in ['admin', 'super_admin']:
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')
    
    month = request.GET.get('month')
    year = request.GET.get('year', datetime.now().year)
    
    if active_role == 'super_admin':
        operators = User.objects.filter(profile__role='operator')
    else:
        operator_ids = request.user.subordinate_operators.values_list('user_id', flat=True)
        operators = User.objects.filter(id__in=operator_ids, profile__role='operator')
    
    ranking = []
    for operator in operators:
        students = Student.objects.filter(created_by=operator)
        
        if month:
            students = students.filter(
                created_at__month=month,
                created_at__year=year
            )
        else:
            students = students.filter(created_at__year=year)
        
        total = students.count()
        approved = students.filter(status='approved').count()
        rejected = students.filter(status='rejected').count()
        pending = students.filter(status='pending').count()
        
        approval_rate = (approved / total * 100) if total > 0 else 0
        
        ranking.append({
            'operator': operator,
            'total': total,
            'approved': approved,
            'rejected': rejected,
            'pending': pending,
            'approval_rate': approval_rate,
            'score': approved * 2 + total * 1
        })
    
    ranking.sort(key=lambda x: x['score'], reverse=True)
    
    for i, item in enumerate(ranking):
        if i == 0:
            item['medal'] = '🥇 Gold'
        elif i == 1:
            item['medal'] = '🥈 Silver'
        elif i == 2:
            item['medal'] = '🥉 Bronze'
        else:
            item['medal'] = None
    
    context = {
        'ranking': ranking,
        'month': month,
        'year': year,
        'months': [
            (1, 'Yanvar'), (2, 'Fevral'), (3, 'Mart'), (4, 'Aprel'),
            (5, 'May'), (6, 'Iyun'), (7, 'Iyul'), (8, 'Avgust'),
            (9, 'Sentyabr'), (10, 'Oktyabr'), (11, 'Noyabr'), (12, 'Dekabr')
        ],
        'unread_notifications': Notification.objects.filter(recipient=request.user, is_read=False).count(),
        'unread_chat_count': get_unread_chat_count(request.user),
    }
    return render(request, 'operator_ranking.html', context)

# --- E'lonlar ---
@login_required
def announcements_list(request):
    # Session dan active_role ni olish
    user_role = request.session.get('active_role', request.user.profile.role)
    can_manage = user_role in ['admin', 'super_admin']
    
    announcements = Announcement.objects.filter(is_active=True)
    
    for ann in announcements:
        if ann.is_expired():
            ann.is_active = False
            ann.save()
    
    announcements = Announcement.objects.filter(is_active=True)
    
    filtered_announcements = []
    for ann in announcements:
        if ann.target_group == 'all':
            filtered_announcements.append(ann)
        elif ann.target_group == user_role:
            filtered_announcements.append(ann)
        elif ann.target_group == 'super_admin' and user_role == 'super_admin':
            filtered_announcements.append(ann)
    
    unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
    
    context = {
        'announcements': filtered_announcements,
        'can_manage': can_manage,
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(request.user),
    }
    return render(request, 'announcements_list.html', context)

@login_required
def create_announcement(request):
    # Session dan active_role ni olish
    active_role = request.session.get('active_role', request.user.profile.role)
    
    if active_role not in ['admin', 'super_admin']:
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('announcements_list')
    
    if request.method == 'POST':
        form = AnnouncementForm(request.POST)
        if form.is_valid():
            announcement = form.save(commit=False)
            announcement.created_by = request.user
            announcement.save()
            
            log_action(request.user, "E'lon yaratildi", announcement.title)
            messages.success(request, "E'lon muvaffaqiyatli yaratildi!")
            return redirect('announcements_list')
    else:
        form = AnnouncementForm()
    
    return render(request, 'create_announcement.html', {
        'form': form,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

@login_required
def delete_announcement(request, pk):
    if request.user.profile.role not in ['admin', 'super_admin']:
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('announcements_list')
    
    announcement = get_object_or_404(Announcement, pk=pk)
    announcement.delete()
    
    messages.success(request, "E'lon o'chirildi!")
    return redirect('announcements_list')

# --- Yangilangan Chat Views ---
@login_required
def chat_list(request):
    """Chat suhbatdoshlari ro'yxati"""
    user_role = request.user.profile.role
    
    all_users = None
    if user_role == 'super_admin':
        all_users = User.objects.exclude(id=request.user.id)
    
    if user_role == 'super_admin':
        sent_messages = ChatMessage.objects.filter(sender=request.user)
        received_messages = ChatMessage.objects.filter(recipient=request.user)
        
        conversations = {}
        
        for msg in sent_messages:
            if msg.recipient.id not in conversations:
                conversations[msg.recipient.id] = {
                    'user': msg.recipient,
                    'last_message': msg,
                    'unread_count': 0
                }
            else:
                if msg.created_at > conversations[msg.recipient.id]['last_message'].created_at:
                    conversations[msg.recipient.id]['last_message'] = msg
        
        for msg in received_messages:
            if msg.sender.id not in conversations:
                conversations[msg.sender.id] = {
                    'user': msg.sender,
                    'last_message': msg,
                    'unread_count': 0
                }
            else:
                if msg.created_at > conversations[msg.sender.id]['last_message'].created_at:
                    conversations[msg.sender.id]['last_message'] = msg
            
            if not msg.is_read:
                conversations[msg.sender.id]['unread_count'] += 1
        
        conversations = list(conversations.values())
        conversations.sort(key=lambda x: x['last_message'].created_at if x['last_message'] else None, reverse=True)
    
    elif user_role == 'operator':
        conversations = {}
        
        if request.user.profile.assigned_admin:
            admin = request.user.profile.assigned_admin
            
            messages_qs = ChatMessage.objects.filter(
                Q(sender=request.user, recipient=admin) |
                Q(sender=admin, recipient=request.user)
            ).order_by('-created_at')
            
            last_message = messages_qs.first() if messages_qs.exists() else None
            
            unread_count = ChatMessage.objects.filter(
                sender=admin,
                recipient=request.user,
                is_read=False
            ).count()
            
            conversations[admin.id] = {
                'user': admin,
                'last_message': last_message,
                'unread_count': unread_count
            }
        
        received_from_others = ChatMessage.objects.filter(
            recipient=request.user
        ).exclude(
            sender=request.user.profile.assigned_admin
        )
        
        for msg in received_from_others:
            if msg.sender.id not in conversations:
                conversations[msg.sender.id] = {
                    'user': msg.sender,
                    'last_message': msg,
                    'unread_count': 0
                }
            
            unread = ChatMessage.objects.filter(
                sender=msg.sender,
                recipient=request.user,
                is_read=False
            ).count()
            
            conversations[msg.sender.id]['unread_count'] = unread
            
            if not conversations[msg.sender.id]['last_message'] or msg.created_at > conversations[msg.sender.id]['last_message'].created_at:
                conversations[msg.sender.id]['last_message'] = msg
        
        sent_messages = ChatMessage.objects.filter(sender=request.user)
        for msg in sent_messages:
            if msg.recipient.id not in conversations:
                conversations[msg.recipient.id] = {
                    'user': msg.recipient,
                    'last_message': msg,
                    'unread_count': 0
                }
            else:
                if msg.created_at > conversations[msg.recipient.id]['last_message'].created_at:
                    conversations[msg.recipient.id]['last_message'] = msg
        
        conversations = list(conversations.values())
        conversations.sort(key=lambda x: x['last_message'].created_at if x['last_message'] else None, reverse=True)
        
        if not conversations:
            messages.info(request, "Hali xabarlar yo'q.")
    
    else:  # admin
        sent_messages = ChatMessage.objects.filter(sender=request.user)
        received_messages = ChatMessage.objects.filter(recipient=request.user)
        
        conversations = {}
        
        for msg in sent_messages:
            if msg.recipient.id not in conversations:
                conversations[msg.recipient.id] = {
                    'user': msg.recipient,
                    'last_message': msg,
                    'unread_count': 0
                }
            else:
                if msg.created_at > conversations[msg.recipient.id]['last_message'].created_at:
                    conversations[msg.recipient.id]['last_message'] = msg
        
        for msg in received_messages:
            if msg.sender.id not in conversations:
                conversations[msg.sender.id] = {
                    'user': msg.sender,
                    'last_message': msg,
                    'unread_count': 0
                }
            else:
                if msg.created_at > conversations[msg.sender.id]['last_message'].created_at:
                    conversations[msg.sender.id]['last_message'] = msg
            
            if not msg.is_read:
                conversations[msg.sender.id]['unread_count'] += 1
        
        conversations = list(conversations.values())
        conversations.sort(key=lambda x: x['last_message'].created_at if x['last_message'] else None, reverse=True)
    
    unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
    
    context = {
        'conversations': conversations,
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(request.user),
        'all_users': all_users,
    }
    return render(request, 'chat_list.html', context)

@login_required
def chat_conversation(request, user_id):
    """Alohida foydalanuvchi bilan chat"""
    other_user = get_object_or_404(User, id=user_id)
    user_role = request.user.profile.role
    
    if user_role == 'operator':
        is_assigned_admin = (request.user.profile.assigned_admin == other_user)
        has_received_message = ChatMessage.objects.filter(
            sender=other_user,
            recipient=request.user
        ).exists()
        has_sent_message = ChatMessage.objects.filter(
            sender=request.user,
            recipient=other_user
        ).exists()
        
        if not (is_assigned_admin or has_received_message or has_sent_message):
            messages.error(request, "Siz bu foydalanuvchi bilan chat qila olmaysiz.")
            return redirect('chat_list')
    
    elif user_role == 'admin':
        if other_user.profile.role == 'operator':
            if other_user.profile.assigned_admin != request.user:
                messages.error(request, "Bu operator sizga biriktirilmagan.")
                return redirect('chat_list')
    
    ChatMessage.objects.filter(
        sender=other_user,
        recipient=request.user,
        is_read=False
    ).update(is_read=True)
    
    messages_qs = ChatMessage.objects.filter(
        Q(sender=request.user, recipient=other_user) |
        Q(sender=other_user, recipient=request.user),
        is_deleted=False
    ).order_by('created_at')
    
    if request.method == 'POST':
        form = ChatMessageForm(request.POST)
        if form.is_valid():
            chat_message = form.save(commit=False)
            chat_message.sender = request.user
            chat_message.recipient = other_user
            
            if chat_message.sticker:
                chat_message.message = chat_message.sticker
            
            chat_message.save()
            return redirect('chat_conversation', user_id=user_id)
    else:
        form = ChatMessageForm()
    
    unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
    
    context = {
        'other_user': other_user,
        'messages': messages_qs,
        'form': form,
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(request.user),
    }
    return render(request, 'chat_conversation.html', context)

# --- Xabarga Reaksiya Qo'shish ---
@login_required
def add_reaction(request, message_id):
    """Xabarga reaksiya qo'shish"""
    if request.method == 'POST':
        message = get_object_or_404(ChatMessage, pk=message_id)
        reaction = request.POST.get('reaction')
        
        if reaction:
            reaction_obj, created = MessageReaction.objects.get_or_create(
                message=message,
                user=request.user,
                defaults={'reaction': reaction}
            )
            
            if not created:
                reaction_obj.reaction = reaction
                reaction_obj.save()
            
            # Faqat super admin audit loglarida ko'rsatish
            if is_super_admin(request.user):
                log_action(request.user, "Reaksiya qo'shildi", 
                          f"Xabar ID: {message_id}, Reaksiya: {reaction}")
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'reaction': reaction})
        
        return redirect('chat_conversation', user_id=message.sender.id if message.recipient == request.user else message.recipient.id)
    
    return redirect('chat_list')

# --- Xabar Reaksiyalarini Olish ---
@login_required
def get_message_reactions(request, message_id):
    """Xabar reaksiyalarini JSON formatda olish"""
    message = get_object_or_404(ChatMessage, pk=message_id)
    
    reactions = []
    for reaction in message.reactions.all():
        reactions.append({
            'reaction': reaction.reaction,
            'user': reaction.user.username
        })
    
    return JsonResponse({'reactions': reactions})

# --- Reaksiyani Olib Tashlash ---
@login_required
def remove_reaction(request, message_id):
    """Xabardan reaksiyani olib tashlash"""
    if request.method == 'POST':
        message = get_object_or_404(ChatMessage, pk=message_id)
        reaction = request.POST.get('reaction')
        
        if reaction:
            try:
                reaction_obj = MessageReaction.objects.get(
                    message=message,
                    user=request.user,
                    reaction=reaction
                )
                reaction_obj.delete()
                
                # Faqat super admin audit loglarida ko'rsatish
                if is_super_admin(request.user):
                    log_action(request.user, "Reaksiya olib tashlandi", 
                              f"Xabar ID: {message_id}, Reaksiya: {reaction}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'removed': True})
                
            except MessageReaction.DoesNotExist:
                pass
        
        return redirect('chat_conversation', user_id=message.sender.id if message.recipient == request.user else message.recipient.id)
    
    return redirect('chat_list')

# --- Xabarni O'chirish (Soft Delete) ---
@login_required
def delete_chat_message(request, pk):
    """Xabarni o'chirish (soft delete)"""
    message = get_object_or_404(ChatMessage, pk=pk, sender=request.user)
    message.is_deleted = True
    message.message = "[Xabar o'chirildi]"
    message.save()
    messages.success(request, "Xabar o'chirildi!")
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    return redirect('chat_conversation', user_id=message.recipient.id)

# --- Xabarni Tahrirlash ---
@login_required
def edit_chat_message(request, pk):
    """Xabarni tahrirlash"""
    message = get_object_or_404(ChatMessage, pk=pk, sender=request.user)
    
    if request.method == 'POST':
        new_message = request.POST.get('message')
        if new_message:
            message.message = new_message
            message.edited_at = timezone.now()
            message.save()
            messages.success(request, "Xabar tahrirlandi!")
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': message.message})
        
        return redirect('chat_conversation', user_id=message.recipient.id)
    
    return redirect('chat_conversation', user_id=message.recipient.id)

# --- Excel Yuklash orqali Talaba Qo'shish ---
@login_required
def upload_students_excel(request):
    """Excel fayli orqali ko'plab talabalarni qo'shish"""
    if request.user.profile.role == 'admin':
        log_action(request.user, 'Admin Excel yuklashga urindi', 'Ruxsat yo\'q')
        messages.error(request, "Adminlar to'g'ridan-to'g'ri talaba qo'sha olmaydi.")
        return redirect('dashboard')

    config = ExcelImportConfig.get_active_config()

    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            # Fayl turlarini tekshirish
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, "Faqat .xlsx yoki .xls formatdagi fayllar yuklash mumkin!")
                return render(request, 'upload_students_excel.html', {'form': form, 'config': config})
            
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active
                
                # Natijalar uchun ro'yxatlar
                added_students = []
                duplicate_passports = []
                failed_students = []
                
                from openpyxl.utils import column_index_from_string
                full_name_idx = column_index_from_string(config.full_name_col) - 1
                passport_idx = column_index_from_string(config.passport_col) - 1
                jshshir_idx = column_index_from_string(config.jshshir_col) - 1
                course_idx = column_index_from_string(config.course_col) - 1
                phone_idx = column_index_from_string(config.phone_col) - 1
                faculty_idx = column_index_from_string(config.faculty_col) - 1
                specialty_idx = column_index_from_string(config.specialty_col) - 1
                operator_comment_idx = column_index_from_string(config.operator_comment_col) - 1

                def get_val(row, idx):
                    if 0 <= idx < len(row):
                        return row[idx]
                    return None
                
                # Qatorlar bo'yicha iteratsiya (2-qatordan boshlab, sarlavhadan keyin)
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):  # Bo'sh qatorlarni o'tkazib yuborish
                        continue
                    
                    try:
                        # Ma'lumotlarni olish (namuna formatiga moslash)
                        full_name = str(get_val(row, full_name_idx)).strip() if get_val(row, full_name_idx) is not None else ""
                        passport = str(get_val(row, passport_idx)).strip().upper() if get_val(row, passport_idx) is not None else ""
                        jshshir = str(get_val(row, jshshir_idx)).strip() if get_val(row, jshshir_idx) is not None else ""
                        course_val = get_val(row, course_idx)
                        course = int(course_val) if course_val and str(course_val).isdigit() else 1
                        phone = str(get_val(row, phone_idx)).strip() if get_val(row, phone_idx) is not None else ""
                        faculty_name = str(get_val(row, faculty_idx)).strip() if get_val(row, faculty_idx) is not None else ""
                        specialty_name = str(get_val(row, specialty_idx)).strip() if get_val(row, specialty_idx) is not None else ""
                        operator_comment = str(get_val(row, operator_comment_idx)).strip() if get_val(row, operator_comment_idx) is not None else ""
                        
                        # Minimal tekshiruvlar
                        if not passport or not full_name:
                            failed_students.append({
                                'row': row_num,
                                'data': row,
                                'reason': 'Pasport yoki to\'liq ism kiritilmagan'
                            })
                            continue
                        
                        # Pasport formatini tekshirish
                        if not re.match(r'^[A-Z]{2}\d{7}$', passport):
                            failed_students.append({
                                'row': row_num,
                                'data': row,
                                'reason': f'Pasport formati noto\'g\'ri: {passport}'
                            })
                            continue
                        
                        # JSHSHIR tekshirish
                        if jshshir and not re.match(r'^\d{14}$', jshshir):
                            failed_students.append({
                                'row': row_num,
                                'data': row,
                                'reason': f'JSHSHIR formati noto\'g\'ri: {jshshir}'
                            })
                            continue
                        
                        # Dublikatni tekshirish
                        if Student.objects.filter(passport=passport).exists():
                            duplicate_passports.append({
                                'row': row_num,
                                'passport': passport,
                                'full_name': full_name
                            })
                            continue
                        
                        # JSHSHIR dublikatni tekshirish
                        if jshshir and Student.objects.filter(jshshir=jshshir).exists():
                            failed_students.append({
                                'row': row_num,
                                'data': row,
                                'reason': f'JSHSHIR bazada mavjud: {jshshir}'
                            })
                            continue
                        
                        # Fakultet va mutaxassislikni tekshirish
                        faculty = None
                        specialty = None
                        
                        if faculty_name:
                            try:
                                faculty = Faculty.objects.get(name=faculty_name)
                            except Faculty.DoesNotExist:
                                failed_students.append({
                                    'row': row_num,
                                    'data': row,
                                    'reason': f"'{faculty_name}' nomli fakultet topilmadi"
                                })
                                continue
                        
                        if specialty_name:
                            try:
                                if faculty:
                                    specialty = Specialty.objects.get(faculty=faculty, name=specialty_name)
                                else:
                                    failed_students.append({
                                        'row': row_num,
                                        'data': row,
                                        'reason': f"'{specialty_name}' mutaxassisligi uchun avval fakultetni to'g'ri kiriting"
                                    })
                                    continue
                            except Specialty.DoesNotExist:
                                failed_students.append({
                                    'row': row_num,
                                    'data': row,
                                    'reason': f"'{specialty_name}' nomli mutaxassislik topilmadi"
                                })
                                continue
                        
                        # Talabani yaratish
                        student = Student.objects.create(
                            passport=passport,
                            jshshir=jshshir,
                            full_name=full_name,
                            course=course,
                            phone=phone,
                            faculty=faculty,
                            specialty=specialty,
                            operator_comment=operator_comment,
                            created_by=request.user
                        )
                        
                        added_students.append({
                            'row': row_num,
                            'passport': passport,
                            'full_name': full_name
                        })
                        
                    except Exception as e:
                        failed_students.append({
                            'row': row_num,
                            'data': row,
                            'reason': f'Xatolik: {str(e)}'
                        })
                
                # Ma'lumotlarni sessionga saqlash
                request.session['failed_students'] = failed_students
                request.session['duplicate_passports'] = duplicate_passports
                
                # Natijalarni log qilish
                log_action(request.user, "Excel yuklandi", 
                          f'Qo\'shildi: {len(added_students)}, Dublikat: {len(duplicate_passports)}, Xatolik: {len(failed_students)}')
                
                # Xabarlar
                if added_students:
                    messages.success(request, f"✅ {len(added_students)} ta talaba muvaffaqiyatli qo'shildi!")
                if duplicate_passports:
                    messages.warning(request, f"⚠️ {len(duplicate_passports)} ta talaba pasport seriyasi takrorlandi!")
                if failed_students:
                    messages.error(request, f"❌ {len(failed_students)} ta talaba qo'shilmadi!")
                
                # Natijalar sahifasiga yo'naltirish
                return render(request, 'upload_results.html', {
                    'added_students': added_students,
                    'duplicate_passports': duplicate_passports,
                    'failed_students': failed_students,
                    'total_rows': len(added_students) + len(duplicate_passports) + len(failed_students),
                    'unread_chat_count': get_unread_chat_count(request.user),
                })
                
            except Exception as e:
                log_action(request.user, "Excel yuklashda xatolik", f'Xatolik: {str(e)}')
                messages.error(request, f"Faylni o'qishda xatolik yuz berdi: {str(e)}")
        else:
            messages.error(request, "Faylni to'g'ri tanlang!")
    else:
        form = ExcelUploadForm()
    
    return render(request, 'upload_students_excel.html', {
        'form': form,
        'config': config,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

# --- FAKULTET VA MUTAXASSISLIKNI EXCEL ORQALI QO'SHISH ---

def _generate_faculty_sample_excel():
    """Fakultetlar uchun namuna Excel faylini yaratish"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fakultetlar"
    
    headers = ["Fakultet Nomi"]
    ws.append(headers)
    
    # Namunalar
    sample_data = [
        ["Kompyuter Texnologiyalari"],
        ["Elektrik Muhandisligi"],
        ["Biolohiya va Kimyo"],
    ]
    for row in sample_data:
        ws.append(row)
    
    # Format
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    ws.column_dimensions['A'].width = 40
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def _generate_specialty_sample_excel():
    """Mutaxassisliklar uchun namuna Excel faylini yaratish"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mutaxassisliklar"
    
    headers = ["Fakultet Nomi", "Mutaxassislik Nomi"]
    ws.append(headers)
    
    # Namunalar
    sample_data = [
        ["Kompyuter Texnologiyalari", "Dasturiy ta'minot"],
        ["Kompyuter Texnologiyalari", "Kompyuter Tarmoqlari"],
        ["Elektrik Muhandisligi", "Elektr Energetikasi"],
        ["Elektrik Muhandisligi", "Elektr Mashinalashtirish"],
    ]
    for row in sample_data:
        ws.append(row)
    
    # Format
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 35
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@login_required
def download_faculty_sample_excel(request):
    """Fakultetlar uchun namuna Excelni yuklab olish"""
    log_action(request.user, "Fakultet namuna Excel yuklab olindi")
    
    buffer = _generate_faculty_sample_excel()
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'fakultetlar_namuna_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

@login_required
def download_specialty_sample_excel(request):
    """Mutaxassisliklar uchun namuna Excelni yuklab olish"""
    log_action(request.user, "Mutaxassislik namuna Excel yuklab olindi")
    
    buffer = _generate_specialty_sample_excel()
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'mutaxassisliklar_namuna_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

@login_required
def upload_faculty_excel(request):
    """Excel fayli orqali ko'plab fakultetlarni qo'shish"""
    if request.user.profile.role not in ['super_admin', 'admin']:
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')

    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, "Faqat .xlsx yoki .xls formatdagi fayllar yuklash mumkin!")
                return render(request, 'upload_faculty_excel.html', {'form': form})
            
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active
                
                added_faculties = []
                duplicate_faculties = []
                failed_faculties = []
                
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):
                        continue
                    
                    try:
                        faculty_name = str(row[0]).strip() if row[0] else ""
                        
                        if not faculty_name:
                            failed_faculties.append({
                                'row': row_num,
                                'reason': 'Fakultet nomi kiritilmagan'
                            })
                            continue
                        
                        if Faculty.objects.filter(name=faculty_name).exists():
                            duplicate_faculties.append({
                                'row': row_num,
                                'name': faculty_name
                            })
                            continue
                        
                        faculty = Faculty.objects.create(name=faculty_name)
                        added_faculties.append({
                            'row': row_num,
                            'name': faculty_name
                        })
                        
                    except Exception as e:
                        failed_faculties.append({
                            'row': row_num,
                            'reason': f'Xatolik: {str(e)}'
                        })
                
                request.session['added_faculties'] = added_faculties
                request.session['duplicate_faculties'] = duplicate_faculties
                request.session['failed_faculties'] = failed_faculties
                
                log_action(request.user, "Fakultetlar Excel yuklandi", 
                          f'Qo\'shildi: {len(added_faculties)}, Dublikat: {len(duplicate_faculties)}, Xatolik: {len(failed_faculties)}')
                
                if added_faculties:
                    messages.success(request, f"✅ {len(added_faculties)} ta fakultet muvaffaqiyatli qo'shildi!")
                if duplicate_faculties:
                    messages.warning(request, f"⚠️ {len(duplicate_faculties)} ta fakultet allaqachon mavjud!")
                if failed_faculties:
                    messages.error(request, f"❌ {len(failed_faculties)} ta fakultet qo'shilmadi!")
                
                return render(request, 'upload_faculty_results.html', {
                    'added_faculties': added_faculties,
                    'duplicate_faculties': duplicate_faculties,
                    'failed_faculties': failed_faculties,
                    'unread_chat_count': get_unread_chat_count(request.user),
                })
                
            except Exception as e:
                log_action(request.user, "Fakultet Excel yuklashda xatolik", f'Xatolik: {str(e)}')
                messages.error(request, f"Faylni o'qishda xatolik yuz berdi: {str(e)}")
        else:
            messages.error(request, "Faylni to'g'ri tanlang!")
    else:
        form = ExcelUploadForm()
    
    return render(request, 'upload_faculty_excel.html', {
        'form': form,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

@login_required
def upload_specialty_excel(request):
    """Excel fayli orqali ko'plab mutaxassisliklarni qo'shish"""
    if request.user.profile.role not in ['super_admin', 'admin']:
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')

    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, "Faqat .xlsx yoki .xls formatdagi fayllar yuklash mumkin!")
                return render(request, 'upload_specialty_excel.html', {'form': form})
            
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active
                
                added_specialties = []
                duplicate_specialties = []
                failed_specialties = []
                
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):
                        continue
                    
                    try:
                        faculty_name = str(row[0]).strip() if row[0] else ""
                        specialty_name = str(row[1]).strip() if row[1] else ""
                        
                        if not faculty_name or not specialty_name:
                            failed_specialties.append({
                                'row': row_num,
                                'reason': 'Fakultet yoki mutaxassislik nomi kiritilmagan'
                            })
                            continue
                        
                        # Fakultetni topish
                        try:
                            faculty = Faculty.objects.get(name=faculty_name)
                        except Faculty.DoesNotExist:
                            failed_specialties.append({
                                'row': row_num,
                                'reason': f"'{faculty_name}' nomli fakultet topilmadi"
                            })
                            continue
                        
                        # Dublikatni tekshirish
                        if Specialty.objects.filter(faculty=faculty, name=specialty_name).exists():
                            duplicate_specialties.append({
                                'row': row_num,
                                'faculty_name': faculty_name,
                                'specialty_name': specialty_name
                            })
                            continue
                        
                        specialty = Specialty.objects.create(
                            faculty=faculty,
                            name=specialty_name
                        )
                        added_specialties.append({
                            'row': row_num,
                            'faculty_name': faculty_name,
                            'specialty_name': specialty_name
                        })
                        
                    except Exception as e:
                        failed_specialties.append({
                            'row': row_num,
                            'reason': f'Xatolik: {str(e)}'
                        })
                
                request.session['added_specialties'] = added_specialties
                request.session['duplicate_specialties'] = duplicate_specialties
                request.session['failed_specialties'] = failed_specialties
                
                log_action(request.user, "Mutaxassisliklar Excel yuklandi", 
                          f'Qo\'shildi: {len(added_specialties)}, Dublikat: {len(duplicate_specialties)}, Xatolik: {len(failed_specialties)}')
                
                if added_specialties:
                    messages.success(request, f"✅ {len(added_specialties)} ta mutaxassislik muvaffaqiyatli qo'shildi!")
                if duplicate_specialties:
                    messages.warning(request, f"⚠️ {len(duplicate_specialties)} ta mutaxassislik allaqachon mavjud!")
                if failed_specialties:
                    messages.error(request, f"❌ {len(failed_specialties)} ta mutaxassislik qo'shilmadi!")
                
                return render(request, 'upload_specialty_results.html', {
                    'added_specialties': added_specialties,
                    'duplicate_specialties': duplicate_specialties,
                    'failed_specialties': failed_specialties,
                    'unread_chat_count': get_unread_chat_count(request.user),
                })
                
            except Exception as e:
                log_action(request.user, "Mutaxassislik Excel yuklashda xatolik", f'Xatolik: {str(e)}')
                messages.error(request, f"Faylni o'qishda xatolik yuz berdi: {str(e)}")
        else:
            messages.error(request, "Faylni to'g'ri tanlang!")
    else:
        form = ExcelUploadForm()
    
    return render(request, 'upload_specialty_excel.html', {
        'form': form,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

# --- Logotip Boshqarish ---
@login_required
def logo_management(request):
    """Logotip boshqarish sahifasi (faqat super admin uchun)"""
    if request.user.profile.role != 'super_admin':
        messages.error(request, "Sizda bu huquq yo'q. Faqat Super Admin logotipni o'zgartira oladi.")
        return redirect('dashboard')
    
    logos = SiteLogo.objects.all()
    active_logo = SiteLogo.get_active_logo()
    
    if request.method == 'POST':
        form = SiteLogoForm(request.POST, request.FILES)
        if form.is_valid():
            logo = form.save()
            log_action(request.user, "Logotip qo'shildi", f'Logotip ID: {logo.id}')
            messages.success(request, "Logotip muvaffaqiyatli qo'shildi!")
            return redirect('logo_management')
    else:
        form = SiteLogoForm()
    
    context = {
        'logos': logos,
        'active_logo': active_logo,
        'form': form,
        'unread_chat_count': get_unread_chat_count(request.user),
    }
    return render(request, 'logo_management.html', context)

@login_required
def toggle_logo(request, pk):
    """Logotipni faol/faolsiz qilish"""
    if request.user.profile.role != 'super_admin':
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')
    
    logo = get_object_or_404(SiteLogo, pk=pk)
    
    # Barcha logotiplarni faolsiz qilish
    SiteLogo.objects.all().update(is_active=False)
    
    # Tanlangan logotipni faol qilish
    logo.is_active = True
    logo.save()
    
    log_action(request.user, "Logotip almashtirildi", f'Logotip ID: {logo.id}')
    messages.success(request, f"Logotip muvaffaqiyatli almashtirildi: {logo}")
    
    return redirect('logo_management')

@login_required
def login_page_settings(request):
    """Kirish sahifasining matn va logo sozlamalari"""
    messages.error(request, "Kirish sahifasi sozlamalari vaqtinchalik faolsizlantirilgan!")
    return redirect('dashboard')

    content = LoginPageContent.get_active_content()
    if request.method == 'POST':
        form = LoginPageContentForm(request.POST, request.FILES, instance=content)
        if form.is_valid():
            saved_content = form.save()
            log_action(request.user, "Kirish sahifasi kontenti yangilandi", f'ID: {saved_content.id}')
            messages.success(request, "Kirish sahifasi ma'lumotlari muvaffaqiyatli saqlandi!")
            return redirect('login_page_settings')
    else:
        form = LoginPageContentForm(instance=content)

    return render(request, 'login_page_settings.html', {
        'form': form,
        'content': content,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

@login_required
def delete_logo(request, pk):
    """Logotipni o'chirish"""
    if request.user.profile.role != 'super_admin':
        messages.error(request, "Sizda bu huquq yo'q.")
        return redirect('dashboard')
    
    logo = get_object_or_404(SiteLogo, pk=pk)
    logo_name = str(logo)
    logo.delete()
    
    log_action(request.user, "Logotip o'chirildi", f'Logotip: {logo_name}')
    messages.success(request, "Logotip muvaffaqiyatli o'chirildi!")
    
    return redirect('logo_management')

# --- Theme Management ---
def live_stream(request, room_name):
    """Jonli efir va real-time chat sahifasi (anonim foydalanuvchilar ham kira oladi)"""
    # Foydalanuvchi tizimga kirmagan bo'lsa, anonim sifatida ishlaydi
    user = request.user
    
    # Session dan active_role ni olish
    if user.is_authenticated:
        active_role = request.session.get('active_role', user.profile.role if hasattr(user, 'profile') else 'operator')
    else:
        active_role = 'viewer'  # Anonim foydalanuvchilar uchun default rol
    
    # Efir boshlanganda barcha foydalanuvchilarga bildirishnoma yuborish
    # Faqat admin va super_admin efir boshlaganda
    if user.is_authenticated and active_role in ['admin', 'super_admin']:
        try:
            from channels.layers import get_channel_layer
            from asgiref.sync import async_to_sync
            
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                "global_notifications",
                {
                    "type": "send_notification",
                    "data": {
                        "action": "live_started",
                        "room_name": room_name,
                        "message": f"Admin {user.username} jonli efir boshladi!"
                    }
                }
            )
            log_action(user, "Jonli efir boshlandi", f"Xona: {room_name}")
        except Exception as e:
            print(f"❌ Global bildirishnoma yuborishda xatolik: {e}")
    
    # Anonim foydalanuvchi uchun xabarlar soni 0
    unread_notifications = 0
    unread_chat_count = 0
    if user.is_authenticated:
        unread_notifications = Notification.objects.filter(recipient=user, is_read=False).count()
        unread_chat_count = get_unread_chat_count(user)
    
    context = {
        'room_name': room_name,
        'active_role': active_role,
        'unread_notifications': unread_notifications,
        'unread_chat_count': unread_chat_count,
    }
    return render(request, 'live_stream.html', context)


@login_required
def save_theme(request):
    """Foydalanuvchi temasini saqlash"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            theme = data.get('theme', 'light')
            
            # Temani foydalanuvchi profilida saqlash
            if hasattr(request.user, 'profile'):
                # Bu yerda temani saqlash mumkin, masalan Profile modelida theme maydoni qo'shish kerak
                pass
            
            return JsonResponse({'status': 'success', 'theme': theme})
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Only POST method allowed'}, status=405)

# --- BUxgalteriya (Accountant) ---
@login_required
def accountant_dashboard(request):
    """BUxgalter uchun dashboard - tasdiqlangan talabalar ro'yxati"""
    messages.error(request, "Buxgalteriya funksiyalari vaqtinchalik faolsizlantirilgan!")
    return redirect('dashboard')
    
    # Faqat tasdiqlangan talabalar
    students = Student.objects.filter(status='approved').select_related('faculty', 'specialty').order_by('-created_at')
    
    # Qidiruv
    search = request.GET.get('search')
    if search:
        students = students.filter(
            Q(full_name__icontains=search) |
            Q(passport__icontains=search) |
            Q(hemis_id__icontains=search)
        )
    
    # Fakultet bo'yicha filter
    faculty_id = request.GET.get('faculty')
    if faculty_id:
        students = students.filter(faculty_id=faculty_id)
    
    unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
    
    return render(request, 'accountant_dashboard.html', {
        'students': students,
        'search': search,
        'faculty_id': faculty_id,
        'faculties': Faculty.objects.all(),
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

@login_required
def add_payment(request, student_id):
    """Talabaga to'lov qo'shish - faqat accountant va super_admin"""
    messages.error(request, "To'lov qo'shish funksiyasi vaqtinchalik faolsizlantirilgan!")
    return redirect('dashboard')
    
    student = get_object_or_404(Student, id=student_id, status='approved')
    
    if request.method == 'POST':
        amount = request.POST.get('amount')
        payment_type = request.POST.get('payment_type', 'tuition')
        payment_date = request.POST.get('payment_date')
        description = request.POST.get('description', '')
        
        if amount and payment_date:
            try:
                from decimal import Decimal
                payment = StudentPayment.objects.create(
                    student=student,
                    amount=Decimal(amount),
                    payment_type=payment_type,
                    payment_date=payment_date,
                    description=description,
                    created_by=request.user
                )
                log_action(request.user, "To'lov qo'shildi", f"{student.full_name} - {amount} so'm")
                messages.success(request, f"{student.full_name} uchun {amount} so'm to'lov muvaffaqiyatli qo'shildi!")
                return redirect('accountant_dashboard')
            except Exception as e:
                messages.error(request, f"Xatolik yuz berdi: {str(e)}")
                import traceback
                print(f"Payment error: {str(e)}")
                print(traceback.format_exc())
        else:
            messages.error(request, "To'lov summasi va sanasi majburiy!")
    
    # Talabaning avvalgi to'lovlari
    payments = student.payments.all().order_by('-payment_date')
    
    # Kontrakt summasi va qolgan qarz
    contract_amount = student.specialty.contract_amount if student.specialty else 0
    total_paid = sum(p.amount for p in payments)
    remaining_debt = contract_amount - total_paid
    
    # Sonni harf ko'rinishida
    from .number_to_words import format_money_uzbek, format_money_short, number_to_words_uzbek
    contract_amount_words = format_money_uzbek(contract_amount)
    total_paid_words = format_money_uzbek(total_paid)
    remaining_debt_words = format_money_uzbek(remaining_debt)
    
    unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
    
    return render(request, 'add_payment.html', {
        'student': student,
        'payments': payments,
        'contract_amount': contract_amount,
        'total_paid': total_paid,
        'remaining_debt': remaining_debt,
        'contract_amount_words': contract_amount_words,
        'total_paid_words': total_paid_words,
        'remaining_debt_words': remaining_debt_words,
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

@login_required
def view_payments(request, student_id):
    """Talabaning to'lovlarini ko'rish - ko'rish huquqi: admin, accountant, super_admin, talaba_viewer, archive"""
    messages.error(request, "To'lovlarni ko'rish funksiyasi vaqtinchalik faolsizlantirilgan!")
    return redirect('dashboard')
    
    student = get_object_or_404(Student, id=student_id, status='approved')
    payments = student.payments.all().order_by('-payment_date')
    
    # Umumiy to'lov
    total_paid = sum(p.amount for p in payments)
    
    # Kontrakt summasi va qolgan qarz
    contract_amount = student.specialty.contract_amount if student.specialty else 0
    remaining_debt = contract_amount - total_paid
    
    # Sonni harf ko'rinishida
    from .number_to_words import format_money_uzbek, format_money_short, number_to_words_uzbek
    contract_amount_words = format_money_uzbek(contract_amount)
    total_paid_words = format_money_uzbek(total_paid)
    remaining_debt_words = format_money_uzbek(remaining_debt)
    
    # To'lov qo'shish huquqi - faqat accountant va super_admin
    user_role = request.session.get('active_role', request.user.profile.role if hasattr(request.user, 'profile') else None)
    can_add_payment = user_role in ['accountant', 'super_admin']
    
    unread_notifications = Notification.objects.filter(recipient=request.user, is_read=False).count()
    
    return render(request, 'view_payments.html', {
        'student': student,
        'payments': payments,
        'total_paid': total_paid,
        'contract_amount': contract_amount,
        'remaining_debt': remaining_debt,
        'contract_amount_words': contract_amount_words,
        'total_paid_words': total_paid_words,
        'remaining_debt_words': remaining_debt_words,
        'can_add_payment': can_add_payment,
        'user_role': user_role,
        'unread_notifications': unread_notifications,
        'unread_chat_count': get_unread_chat_count(request.user),
    })

# --- API Token ---
@api_view(['POST'])
@permission_classes([AllowAny])
def api_token_auth(request):
    username = request.data.get('username')
    password = request.data.get('password')
    
    user = authenticate(username=username, password=password)
    
    if user:
        token, created = Token.objects.get_or_create(user=user)
        return JsonResponse({
            'token': token.key,
            'user_id': user.pk,
            'username': user.username,
            'role': user.profile.role if hasattr(user, 'profile') else 'operator'
        })
    else:
        return JsonResponse({'error': 'Invalid credentials'}, status=400)




                                                     
