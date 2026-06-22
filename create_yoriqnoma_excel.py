#!/usr/bin/env python3
"""
Yo'riqnoma Excel faylini yaratish uchun skript
"""

import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def create_yoriqnoma_excel():
    """Yo'riqnoma ma'lumotlari uchun Excel faylini yaratish"""
    
    # Aniq formatdagi namuna ma'lumotlar
    data = {
        'To\'liq ism familiya': [
            'Aliyev Ali Vali o\'g\'li',
            'Karimova Karima O\'tkir qizi',
            'To\'chiev To\'chirbek Bahodir o\'g\'li',
            'Nazarova Nazira Rustam qizi',
            'Saidova Saida Aziz qizi'
        ],
        'Pasport seriya va raqami (AA1234567)': [
            'AA1234567',
            'BB2345678',
            'CC3456789',
            'DD4567890',
            'EE5678901'
        ],
        'JSHSHIR (14 ta raqam)': [
            '12345678901234',
            '23456789012345',
            '34567890123456',
            '45678901234567',
            '56789012345678'
        ],
        'Kurs (1, 2, 3, 4)': [
            '3',
            '2',
            '4',
            '1',
            '2'
        ],
        'Telefon raqam': [
            '998901234567',
            '998902345678',
            '998903456789',
            '998904567890',
            '998905678901'
        ],
        'Fakultet nomi': [
            'Informatika va axborot texnologiyalari fakulteti',
            'Matematika fakulteti',
            'Fizika fakulteti',
            'Kimyo fakulteti',
            'Iqtisodiyot fakulteti'
        ],
        'Mutaxassislik nomi': [
            'Komp\'yuter injiniring',
            'Amaliy matematika',
            'Nazariy fizika',
            'Organik kimyo',
            'Moliya va bank ishi'
        ],
        'Operator izohi (ixtiyoriy)': [
            'A\'lo talaba',
            'Yaxshi natijalar',
            'Ilmiy ishtirokchi',
            'Aktiv talaba',
            'Ijtimoiy faol'
        ]
    }
    
    # DataFrame yaratish
    df = pd.DataFrame(data)
    
    # Fayl yo'lini belgilash
    media_dir = '/home/riva/mini_hemis/media'
    if not os.path.exists(media_dir):
        os.makedirs(media_dir)
    
    file_path = os.path.join(media_dir, 'yoriqnoma_namuna.xlsx')
    
    # Excel faylini saqlash
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Yo\'riqnoma', index=False)
        
        # Formatlash
        workbook = writer.book
        worksheet = writer.sheets['Yo\'riqnoma']
        
        # Sarlavha qatorini formatlash
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ustun kengliklarini o'rnatish
        column_widths = {
            'A': 30,  # To'liq ism familiya
            'B': 25,  # Pasport
            'C': 20,  # JSHSHIR
            'D': 15,  # Kurs
            'E': 18,  # Telefon
            'F': 35,  # Fakultet
            'G': 30,  # Mutaxassislik
            'H': 25   # Operator izohi
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Barcha kataklarni formatlash
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.row > 1:  # Sarlavha emas
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.font = Font(size=11)
    
    print(f"Yo'riqnoma Excel fayli yaratildi: {file_path}")
    print(f"Jami {len(df)} ta talaba ma'lumoti qo'shildi")
    print("\nUstunlar:")
    for i, col in enumerate(df.columns, 1):
        print(f"{i}. {col}")
    
    return file_path

if __name__ == "__main__":
    create_yoriqnoma_excel()
